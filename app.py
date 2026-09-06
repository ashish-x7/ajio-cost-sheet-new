from __future__ import annotations

import html
import io
import json
import os
import re
import sys
from typing import Iterable

BASE_DIR = os.path.dirname(__file__)
sys.path.insert(0, os.path.join(BASE_DIR, ".venv_deps"))
sys.path.insert(0, os.path.join(os.path.dirname(BASE_DIR), ".venv_deps"))

import requests
from flask import Flask, Response, jsonify, request, send_from_directory
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from python_calamine import CalamineWorkbook
import xlrd


try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

if load_dotenv:
    load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB max request size

GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"


EXPECTED_HEADERS = [
    "No.",
    "Seller Name",
    "Item Name",
    "Category",
    "Brand",
    "Type",
    "ASIN Number",
    "SKU Number",
    "HSN Number",
    "MRP Price",
    "Item Color",
    "Weight",
    "Weight Unit",
    "Length",
    "Length Unit",
    "Width",
    "Width Unit",
    "Height",
    "Height Unit",
    "Channel Price",
    "Purchase Margin(%)",
    "Tax",
    "Purchase Cost",
    "Purchase Tax",
    "Final Purchase Cost",
    "JIO CODE",
    "COMPANY PROFIT MARGIN %",
    "SALE TP WITH OUT TAX",
    "GST%",
    "SALE TP WITH TAX",
    "MRP ",
    "SALE DISCOUNT AMT",
    "ASP (GROSS)",
    "GST%",
    "GST amount",
    "NET SALE",
    "AJIO MARGIN 34%",
    "PURCHASE",
    "GST%2",
    "GST2 amount",
    "BANK SETTLEMENT",
]


def normalize_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def find_header_index(rows: list[list[object]]) -> int:
    expected = {header.lower() for header in EXPECTED_HEADERS}
    best_index = 0
    best_score = -1

    for index, row in enumerate(rows[:15]):
        lowered = {
            str(cell).lower()
            for cell in row
            if cell is not None and str(cell).strip()
        }
        score = len(expected.intersection(lowered))
        if score > best_score:
            best_score = score
            best_index = index

    return best_index


def parse_excel(file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    raw_rows = [[normalize_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    return finalize_rows(raw_rows)


def parse_xls(file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)
    raw_rows = []

    for row_index in range(sheet.nrows):
        row = [normalize_cell(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
        raw_rows.append(row)

    return finalize_rows(raw_rows)


def parse_with_calamine(file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = CalamineWorkbook.from_filelike(io.BytesIO(file_bytes))
    sheet = workbook.get_sheet_by_index(0)
    raw_rows = []

    for row in sheet.to_python(skip_empty_area=False):
        raw_rows.append([normalize_cell(cell) for cell in row])

    return finalize_rows(raw_rows)


def finalize_rows(raw_rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    while raw_rows and not any(raw_rows[0]):
        raw_rows.pop(0)

    if not raw_rows:
        return EXPECTED_HEADERS, []

    header_index = find_header_index(raw_rows)
    header = raw_rows[header_index]
    data_rows = [row for row in raw_rows[header_index + 1 :] if any(cell != "" for cell in row)]
    column_count = max(len(header), *(len(row) for row in data_rows), len(EXPECTED_HEADERS))
    header = header + [""] * (column_count - len(header))
    for i, expected in enumerate(EXPECTED_HEADERS):
        if i < len(header) and not header[i]:
            header[i] = expected
    data_rows = [row + [""] * (column_count - len(row)) for row in data_rows]
    return header, data_rows


def extract_uploaded_file(content_type: str, body: bytes) -> tuple[str, bytes]:
    if "boundary=" not in content_type:
        raise ValueError("Invalid upload request.")

    boundary = content_type.split("boundary=", 1)[1].encode("utf-8")
    delimiter = b"--" + boundary

    for part in body.split(delimiter):
        if b"filename=" not in part:
            continue

        try:
            header_block, file_block = part.split(b"\r\n\r\n", 1)
        except ValueError as exc:
            raise ValueError("Upload format samajh nahi aaya.") from exc

        headers = header_block.decode("utf-8", errors="ignore")
        filename = "uploaded.xlsx"
        for line in headers.split("\r\n"):
            if "filename=" in line:
                filename = line.split("filename=", 1)[1].strip().strip('"')
                break

        file_bytes = file_block.rsplit(b"\r\n", 1)[0]
        return filename, file_bytes

    raise ValueError("Koi file receive nahi hui.")


def get_col_letter(col_idx: int) -> str:
    result = ""
    while col_idx >= 0:
        result = chr(col_idx % 26 + 65) + result
        col_idx = col_idx // 26 - 1
    return result


def render_table(headers: Iterable[str], rows: Iterable[Iterable[str]]) -> str:
    headers_list = list(headers)
    display_headers = [f"{get_col_letter(i)} - {cell}" for i, cell in enumerate(headers_list)]
    header_html = "".join(f"<th>{html.escape(cell)}</th>" for cell in display_headers)
    
    rows_list = list(rows)
    rows_json = json.dumps(rows_list).replace("</", "<\\/")
    headers_json = json.dumps(headers_list).replace("</", "<\\/")
    colspan = max(1, len(headers_list))
    
    empty_msg = f"<tr><td colspan='{colspan}' class='empty'>Upload ke baad rows yahan dikhengi.</td></tr>" if not rows_list else ""

    return (
        "<div class='table-wrap'>"
        "<table>"
        f"<thead><tr>{header_html}</tr></thead>"
        f"<tbody id='table-body'>{empty_msg}</tbody>"
        "</table>"
        "</div>"
        f"<script>const globalTableData = {rows_json};</script>"
        f"<script>const globalHeaders = {headers_json};</script>"
    )


def sanitize_filename(filename: str) -> str:
    cleaned = "".join(char for char in filename if char not in '<>:"/\\|?*').strip().strip(".")
    return cleaned or "ajio_export"


def coerce_export_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        text = ILLEGAL_CHARACTERS_RE.sub("", value).strip()
        if not text:
            return ""
        if text.startswith("="):
            return text
        cleaned = text.replace(",", "")
        is_percent = cleaned.endswith("%")
        if is_percent:
            cleaned = cleaned[:-1].strip()
        if re.fullmatch(r"[+-]?\d+", cleaned):
            try:
                num = int(cleaned)
                return float(num) / 100.0 if is_percent else num
            except ValueError:
                return text
        if re.fullmatch(r"[+-]?(?:\d*\.\d+|\d+\.\d*)", cleaned):
            try:
                num = float(cleaned)
                return num / 100.0 if is_percent else num
            except ValueError:
                return text
        return text
    return value


def build_export_workbook(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AJIO Export"

    sheet.append(headers)
    for row_index, row in enumerate(rows, start=2):
        padded = row + [""] * (len(headers) - len(row))
        coerced_row = [coerce_export_cell(cell) for cell in padded[: len(headers)]]
        for col_idx, cell_value in enumerate(coerced_row, start=1):
            cell = sheet.cell(row=row_index, column=col_idx, value=cell_value)
            if (
                isinstance(cell_value, (int, float))
                and "%" in headers[col_idx - 1]
                and abs(cell_value) <= 1
            ):
                cell.number_format = "0%"

    sheet.freeze_panes = "A2"

    red_fill = PatternFill(fill_type="solid", fgColor="FDE2E2")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF4BF")
    green_fill = PatternFill(fill_type="solid", fgColor="DDF6E4")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Apply styles to header row only (including borders)
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        if col_idx <= 25:
            cell.fill = red_fill
        elif col_idx == 26:
            cell.fill = yellow_fill
        else:
            cell.fill = green_fill

    # NO borders for data rows (as requested)
    if sheet.max_row > 1:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    # Calculate column widths efficiently
    column_widths = {}
    
    # Get header widths
    for col_idx in range(1, len(headers) + 1):
        header_val = str(headers[col_idx - 1] or "")
        column_widths[col_idx] = max(len(header_val) + 3, 12)

    # Sample data rows for width - strategic sampling for large files
    if len(rows) > 100:
        sample_size = min(500, max(100, len(rows) // 10))
        sample_indices = [int(i * len(rows) / sample_size) for i in range(sample_size)]
    else:
        sample_indices = range(len(rows))
    
    for row_idx in sample_indices:
        if row_idx < len(rows):
            for col_idx in range(1, len(headers) + 1):
                if col_idx - 1 < len(rows[row_idx]):
                    cell_val = str(rows[row_idx][col_idx - 1] or "")
                    if col_idx == 3:
                        # Item Name column - special handling
                        width = min(max(len(cell_val) + 3, 32), 80)
                    else:
                        width = min(max(len(cell_val) + 3, 12), 28)
                    column_widths[col_idx] = max(column_widths.get(col_idx, 12), width)

    # Apply column widths
    for col_idx, width in column_widths.items():
        column_letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[column_letter].width = width

    # Set row heights only for header
    sheet.row_dimensions[1].height = 20

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_page(table_html: str = "", message: str = "", default_tab: str = "auto") -> bytes:
    message_html = f"<p class='message'>{html.escape(message)}</p>" if message else "<p class='message'></p>"
    if not table_html:
        table_html = render_table(EXPECTED_HEADERS, [])

    page = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>AJIO Cost calculator</title>
  <link rel="icon" type="image/png" href="/logo.png">
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:opsz,wght@14..32,400;14..32,600;14..32,700;14..32,800;14..32,900&display=swap');
    :root {{
      --bg: #f8fafc;
      --surface: #ffffff;
      --surface-soft: #f1f5f9;
      --border: #e2e8f0;
      --border-strong: #cbd5e1;
      --text: #000814;
      --muted: #1e293b;
      --muted-soft: #475569;
      --slate: #334155;
      --emerald: #059669;
      --emerald-soft: #ecfdf5;
      --rose: #e11d48;
      --rose-soft: #fff1f2;
      --amber: #b45309;
      --amber-soft: #fffbeb;
      --shadow-sm: 0 1px 2px rgba(15, 23, 42, 0.06);
      --shadow-md: 0 14px 32px rgba(15, 23, 42, 0.08);
      --shadow-xl: 0 24px 70px rgba(15, 23, 42, 0.18);
    }}
    * {{ box-sizing: border-box; }}
    html {{ background: var(--bg); }}
    body {{
      margin: 0;
      min-height: 100vh;
      overflow-x: hidden;
      background:
        linear-gradient(180deg, rgba(255,255,255,0.94), rgba(248,250,252,0.98) 260px),
        var(--bg);
      color: var(--text);
      font-family: "Inter", "Segoe UI", Arial, sans-serif;
      font-size: 14px;
      line-height: 1.5;
    }}
    button, input, select, textarea {{ font: inherit; }}
    .shell {{
      width: 100%;
      max-width: none;
      min-height: 100vh;
      margin: 0;
      padding: 16px;
    }}
    .hero {{
      display: flex;
      align-items: center;
      gap: 14px;
      margin-bottom: 16px;
      padding: 18px;
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 24px;
      box-shadow: var(--shadow-sm);
    }}
    .brand-logo {{
      width: 54px;
      height: 54px;
      object-fit: contain;
      padding: 8px;
      flex-shrink: 0;
      background: var(--surface-soft);
      border: 1px solid var(--border);
      border-radius: 18px;
    }}
    h1 {{
      margin: 0;
      color: var(--text);
      font-size: clamp(24px, 3vw, 40px);
      font-weight: 900;
      font-style: italic;
      letter-spacing: -0.055em;
      line-height: 0.95;
    }}
    .myntra-btn {{
      margin-left: auto;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      min-height: 42px;
      padding: 0 14px;
      color: #be123c;
      background: var(--rose-soft);
      border: 1px solid #fecdd3;
      border-radius: 14px;
      text-decoration: none;
      font-size: 11px;
      font-weight: 900;
      letter-spacing: 0.12em;
      text-transform: uppercase;
      box-shadow: var(--shadow-sm);
      transition: all 180ms ease;
    }}
    .myntra-btn:hover {{ transform: translateY(-1px); box-shadow: var(--shadow-md); }}
    form {{
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 12px;
      margin: 0;
      padding: 0;
    }}
    .global-inputs, .tabs-container, .upload-zone, #single-container > div:first-child {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 22px;
      box-shadow: var(--shadow-sm);
    }}
    .global-inputs {{
      display: flex;
      flex-wrap: wrap;
      align-items: flex-start;
      gap: 10px;
      margin-bottom: 16px;
      padding: 10px;
      width: 900px;
      max-width: 100%;
    }}
    .margin-input-group {{
      display: flex;
      flex-direction: column;
      align-items: stretch;
      gap: 7px;
      width: 210px;
      min-width: 180px;
      padding: 10px;
      background: #ffffff !important;
      border: 1px solid var(--border);
      border-radius: 18px;
      box-shadow: var(--shadow-sm);
      transition: all 220ms ease;
    }}
    .margin-input-group:hover {{
      transform: translateY(-1px);
      border-color: var(--border-strong);
      box-shadow: var(--shadow-md);
    }}
    .global-inputs .margin-input-group:nth-child(1) {{
      background: #fff1f2 !important;
      border-color: #fecdd3;
    }}
    .global-inputs .margin-input-group:nth-child(1) label {{
      color: #9f1239 !important;
      font-weight: 400 !important;
    }}
    .global-inputs .margin-input-group:nth-child(1) input {{
      border-color: #fecdd3 !important;
    }}
    .global-inputs .margin-input-group:nth-child(2) {{
      background: #eff6ff !important;
      border-color: #bfdbfe;
    }}
    .global-inputs .margin-input-group:nth-child(2) label {{
      color: #1e40af !important;
      font-weight: 400 !important;
    }}
    .global-inputs .margin-input-group:nth-child(2) input {{
      border-color: #bfdbfe !important;
    }}
    .global-inputs .margin-input-group:nth-child(3) {{
      background: #ecfdf5 !important;
      border-color: #a7f3d0;
    }}
    .global-inputs .margin-input-group:nth-child(3) label {{
      color: #065f46 !important;
      font-weight: 400 !important;
    }}
    .global-inputs .margin-input-group:nth-child(3) input {{
      border-color: #a7f3d0 !important;
    }}
    .global-inputs .margin-input-group:nth-child(4) {{
      background: #fff7ed !important;
      border-color: #fed7aa;
    }}
    .global-inputs .margin-input-group:nth-child(4) label {{
      color: #9a3412 !important;
      font-weight: 400 !important;
    }}
    .global-inputs .margin-input-group:nth-child(4) input {{
      border-color: #fed7aa !important;
    }}
    #single-container > div:first-child .margin-input-group:nth-child(1) {{
      background: #fefce8 !important;
      border-color: #fde68a;
    }}
    #single-container > div:first-child .margin-input-group:nth-child(1) label {{
      color: #854d0e !important;
      font-weight: 400 !important;
    }}
    #single-container > div:first-child .margin-input-group:nth-child(1) input {{
      border-color: #fde68a !important;
    }}
    #single-container > div:first-child .margin-input-group:nth-child(2) {{
      background: #faf5ff !important;
      border-color: #e9d5ff;
    }}
    #single-container > div:first-child .margin-input-group:nth-child(2) label {{
      color: #6b21a8 !important;
      font-weight: 400 !important;
    }}
    #single-container > div:first-child .margin-input-group:nth-child(2) input {{
      border-color: #e9d5ff !important;
    }}
    .margin-input-group label, .toggle-label, #pagination-controls label, #pagination-controls span {{
      color: var(--muted) !important;
      font-size: 9px !important;
      font-weight: 900 !important;
      letter-spacing: 0.13em !important;
      line-height: 1.2;
      text-transform: uppercase;
    }}
    .margin-input-group input, #export-file-name, .bot-input-area input, select {{
      width: 100%;
      min-height: 36px;
      padding: 7px 10px !important;
      color: var(--text) !important;
      background: #ffffff !important;
      border: 1px solid var(--border) !important;
      border-radius: 14px !important;
      outline: none;
      font-size: 14px !important;
      font-weight: 700 !important;
      box-shadow: var(--shadow-sm) !important;
      transition: all 180ms ease;
    }}
    .margin-input-group input:focus, #export-file-name:focus, .bot-input-area input:focus, select:focus {{
      border-color: #94a3b8 !important;
      box-shadow: 0 0 0 3px rgba(148, 163, 184, 0.25) !important;
    }}
    .margin-input-group input::placeholder, #export-file-name::placeholder, .bot-input-area input::placeholder {{
      color: var(--muted-soft);
    }}
    .tabs-container {{
      display: flex;
      gap: 8px;
      margin-bottom: 16px;
      padding: 8px;
      width: fit-content;
      max-width: 100%;
    }}
    .tab-btn {{
      flex: 0 0 auto;
      min-height: 34px;
      padding: 0 14px !important;
      color: var(--muted);
      background: transparent;
      border: 1px solid transparent;
      border-radius: 16px;
      box-shadow: none;
      cursor: pointer;
      font-size: 11px;
      font-weight: 900;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      transition: all 180ms ease;
    }}
    .tab-btn:hover {{
      color: var(--text);
      background: var(--surface-soft);
      transform: translateY(-1px);
    }}
    .tab-btn.active {{
      color: #ffffff;
      background: var(--slate);
      border-color: var(--slate);
      box-shadow: 0 10px 24px rgba(51, 65, 85, 0.22);
    }}
    #single-container > div:first-child {{
      margin-bottom: 16px;
      padding: 14px !important;
    }}
    .action-group {{
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 10px;
    }}
    button {{
      border: 1px solid var(--border);
      border-radius: 14px;
      background: #ffffff;
      color: var(--text);
      cursor: pointer;
      box-shadow: var(--shadow-sm);
      transition: all 180ms ease;
    }}
    button:hover {{ transform: translateY(-1px); box-shadow: var(--shadow-md); }}
    button:disabled {{ opacity: 0.55; cursor: not-allowed; transform: none; }}
    .primary-btn {{
      display: inline-flex !important;
      align-items: center !important;
      justify-content: center !important;
      gap: 7px !important;
      min-height: 42px !important;
      padding: 0 16px !important;
      color: #ffffff !important;
      background: var(--slate) !important;
      border: 1px solid var(--slate) !important;
      border-radius: 14px !important;
      font-size: 11px !important;
      font-weight: 900 !important;
      letter-spacing: 0.08em !important;
      text-transform: uppercase !important;
      white-space: nowrap;
      box-shadow: 0 10px 22px rgba(51, 65, 85, 0.18) !important;
    }}
    .primary-btn:hover {{ box-shadow: var(--shadow-md) !important; }}
    .btn-export {{
      color: #047857 !important;
      background: var(--emerald-soft) !important;
      border-color: #a7f3d0 !important;
      box-shadow: var(--shadow-sm) !important;
    }}
    .btn-clear {{
      color: var(--rose) !important;
      background: var(--rose-soft) !important;
      border-color: #fecdd3 !important;
      box-shadow: var(--shadow-sm) !important;
    }}
    .btn-back {{
      color: var(--muted) !important;
      background: var(--surface-soft) !important;
      border-color: var(--border) !important;
      box-shadow: var(--shadow-sm) !important;
    }}
    .upload-zone {{
      display: flex;
      align-items: center;
      justify-content: flex-start;
      flex-wrap: wrap;
      gap: 14px;
      margin-bottom: 16px;
      padding: 10px;
      width: fit-content;
      max-width: calc(100vw - 32px);
    }}
    #bulk-upload-form {{
      flex: 0 0 auto;
      display: flex;
      align-items: center;
      flex-wrap: wrap;
      gap: 12px;
      min-width: 0;
      max-width: 100%;
    }}
    .file-input-wrapper {{
      position: relative;
      flex: 0 0 auto;
      width: 210px;
      min-width: 180px;
      max-width: 260px !important;
    }}
    #bulk-file-input {{
      position: absolute;
      width: 1px;
      height: 1px;
      opacity: 0;
      overflow: hidden;
      z-index: -1;
    }}
    .file-label {{
      display: flex;
      align-items: center;
      gap: 10px;
      min-height: 36px;
      width: 100%;
      padding: 6px 12px !important;
      color: var(--text);
      background: #ffffff;
      border: 1px dashed var(--border-strong);
      border-radius: 14px;
      cursor: pointer;
      box-shadow: var(--shadow-sm);
      transition: all 180ms ease;
    }}
    .file-label:hover {{
      border-color: #94a3b8;
      background: #f8fafc;
    }}
    .file-label .icon {{ font-size: 15px; line-height: 1; }}
    .file-label .text {{
      min-width: 0;
      overflow: hidden;
      color: var(--text);
      font-size: 12px;
      font-weight: 800;
      text-overflow: ellipsis;
      white-space: nowrap;
    }}
    .sample-link {{
      color: var(--muted);
      font-size: 10px;
      font-weight: 900;
      letter-spacing: 0.1em;
      text-decoration: none;
      text-transform: uppercase;
      border-bottom: 1px solid var(--border-strong);
      transition: all 180ms ease;
    }}
    .sample-link:hover {{ color: var(--text); border-color: var(--text); }}
    .toggle-container {{
      display: inline-flex;
      align-items: center;
      gap: 10px;
      min-height: 36px;
      padding: 6px 10px;
      background: #ffffff;
      border: 1px solid var(--border);
      border-radius: 14px;
      box-shadow: var(--shadow-sm);
      cursor: pointer;
      user-select: none;
    }}
    .switch {{ position: relative; display: inline-block; width: 42px; height: 24px; flex: 0 0 auto; }}
    .switch input {{ width: 0; height: 0; opacity: 0; }}
    .slider {{
      position: absolute;
      inset: 0;
      background: #cbd5e1;
      border-radius: 999px;
      cursor: pointer;
      transition: all 180ms ease;
    }}
    .slider:before {{
      content: "";
      position: absolute;
      width: 18px;
      height: 18px;
      left: 3px;
      top: 3px;
      background: #ffffff;
      border-radius: 999px;
      box-shadow: 0 2px 8px rgba(15, 23, 42, 0.18);
      transition: all 180ms ease;
    }}
    input:checked + .slider {{ background: #10b981; }}
    input:checked + .slider:before {{ transform: translateX(18px); }}
    .file-info {{
      display: inline-flex;
      align-items: center;
      min-height: 36px;
      margin: 0 0 14px;
      padding: 8px 12px;
      color: var(--emerald);
      background: var(--emerald-soft);
      border: 1px solid #a7f3d0;
      border-radius: 14px;
      font-size: 11px;
      font-weight: 800;
    }}
    .single-results-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 10px;
      padding: 0 0 24px;
    }}
    .result-card {{
      display: flex;
      flex-direction: column;
      justify-content: space-between;
      gap: 12px;
      min-height: 92px;
      padding: 12px;
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 20px;
      box-shadow: var(--shadow-sm);
      transition: all 220ms ease;
    }}
    .result-card:hover {{
      transform: scale(1.01);
      border-color: var(--border-strong);
      box-shadow: var(--shadow-md);
    }}
    .result-card h4 {{
      margin: 0;
      color: var(--muted);
      font-size: 8px;
      font-weight: 900;
      letter-spacing: 0.13em;
      line-height: 1.35;
      text-transform: uppercase;
    }}
    .result-card p {{
      margin: 0;
      color: var(--text);
      font-size: 24px;
      font-style: italic;
      font-weight: 900;
      letter-spacing: -0.055em;
      line-height: 0.95;
      text-align: right;
    }}
    .highlight-card {{
      background: linear-gradient(180deg, #ffffff, #f8fafc);
      border-color: #a7f3d0;
      box-shadow: 0 16px 38px rgba(5, 150, 105, 0.09);
    }}
    .highlight-card p {{ color: var(--emerald); }}
    .single-results-grid .result-card:nth-child(1),
    .single-results-grid .result-card:nth-child(2),
    .single-results-grid .result-card:nth-child(12) {{
      background: #eff6ff;
      border-color: #bfdbfe;
    }}
    .single-results-grid .result-card:nth-child(1) h4,
    .single-results-grid .result-card:nth-child(2) h4,
    .single-results-grid .result-card:nth-child(12) h4 {{
      color: #1e40af !important;
    }}
    .single-results-grid .result-card:nth-child(3),
    .single-results-grid .result-card:nth-child(11),
    .single-results-grid .result-card:nth-child(15) {{
      background: #ecfdf5;
      border-color: #a7f3d0;
    }}
    .single-results-grid .result-card:nth-child(3) h4,
    .single-results-grid .result-card:nth-child(11) h4,
    .single-results-grid .result-card:nth-child(15) h4 {{
      color: #065f46 !important;
    }}
    .single-results-grid .result-card:nth-child(4),
    .single-results-grid .result-card:nth-child(5),
    .single-results-grid .result-card:nth-child(7),
    .single-results-grid .result-card:nth-child(10) {{
      background: #fff7ed;
      border-color: #fed7aa;
    }}
    .single-results-grid .result-card:nth-child(4) h4,
    .single-results-grid .result-card:nth-child(5) h4,
    .single-results-grid .result-card:nth-child(7) h4,
    .single-results-grid .result-card:nth-child(10) h4 {{
      color: #9a3412 !important;
    }}
    .single-results-grid .result-card:nth-child(6) {{
      background: #fff1f2;
      border-color: #fecdd3;
    }}
    .single-results-grid .result-card:nth-child(6) h4 {{
      color: #9f1239 !important;
    }}
    .single-results-grid .result-card:nth-child(8),
    .single-results-grid .result-card:nth-child(9),
    .single-results-grid .result-card:nth-child(13),
    .single-results-grid .result-card:nth-child(14) {{
      background: #eef2ff;
      border-color: #c7d2fe;
    }}
    .single-results-grid .result-card:nth-child(8) h4,
    .single-results-grid .result-card:nth-child(9) h4,
    .single-results-grid .result-card:nth-child(13) h4,
    .single-results-grid .result-card:nth-child(14) h4 {{
      color: #3730a3 !important;
    }}
    .table-wrap {{
      width: 100%;
      max-width: 100%;
      min-width: 0;
      overflow: auto;
      background: #ffffff;
      border: 1px solid var(--border);
      border-radius: 22px;
      box-shadow: var(--shadow-sm);
    }}
    .table-wrap .table-wrap {{
      border: 0;
      border-radius: 0;
      box-shadow: none;
    }}
    table {{
      width: 100%;
      min-width: 1260px;
      border-collapse: separate;
      border-spacing: 0;
      background: #ffffff;
    }}
    th, td {{
      padding: 11px 12px;
      border-bottom: 1px solid var(--border);
      border-right: 1px solid #f1f5f9;
      text-align: left;
      vertical-align: top;
      white-space: nowrap;
    }}
    th {{
      position: sticky;
      top: 0;
      z-index: 2;
      color: var(--muted);
      background: #f8fafc;
      font-size: 9px;
      font-weight: 900;
      letter-spacing: 0.13em;
      line-height: 1.35;
      text-transform: uppercase;
    }}
    td {{
      color: #334155;
      font-size: 12px;
      font-weight: 600;
    }}
    tbody tr:hover td {{ background: #f8fafc; }}
    .empty {{
      padding: 40px;
      color: var(--muted);
      text-align: center;
      font-size: 12px;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
    }}
    #pagination-controls {{
      background: #ffffff !important;
      border: 1px solid var(--border) !important;
      border-radius: 18px !important;
      box-shadow: var(--shadow-sm);
      backdrop-filter: none !important;
    }}
    .page-btn {{
      min-height: 36px;
      padding: 0 14px !important;
      color: var(--text);
      background: #ffffff;
      border: 1px solid var(--border);
      border-radius: 12px !important;
      font-size: 10px;
      font-weight: 900 !important;
      letter-spacing: 0.1em;
      text-transform: uppercase;
    }}
    .popup-backdrop {{
      position: fixed;
      inset: 0;
      z-index: 50;
      display: none;
      align-items: center;
      justify-content: center;
      padding: 16px;
      background: rgba(15, 23, 42, 0.4);
      backdrop-filter: blur(4px);
      -webkit-backdrop-filter: blur(4px);
    }}
    .popup-backdrop.open {{ display: flex; }}
    .popup-card {{
      width: min(420px, 100%);
      padding: 24px;
      background: #ffffff;
      border: 1px solid var(--border);
      border-radius: 20px;
      box-shadow: var(--shadow-xl);
      animation: modal-in 150ms ease both;
    }}
    @keyframes modal-in {{
      from {{ opacity: 0; transform: scale(0.95) translateY(6px); }}
      to {{ opacity: 1; transform: scale(1) translateY(0); }}
    }}
    .popup-card h3 {{
      margin: 0 0 8px !important;
      color: var(--text) !important;
      font-size: 18px !important;
      font-weight: 900 !important;
      letter-spacing: -0.03em !important;
    }}
    .popup-card p {{
      margin: 0 0 18px !important;
      color: var(--muted) !important;
      font-size: 13px !important;
      line-height: 1.6 !important;
    }}
    #export-file-name {{ margin-bottom: 16px !important; }}
    .popup-actions {{
      display: flex;
      justify-content: flex-end;
      gap: 10px;
      margin-top: 18px;
    }}
    .bot-trigger {{
      position: fixed;
      right: 24px;
      bottom: 24px;
      z-index: 2000;
      width: 58px;
      height: 58px;
      padding: 0;
      background: transparent;
      border: 0;
      border-radius: 0;
      box-shadow: none;
      cursor: pointer;
      transition: all 220ms ease;
    }}
    .bot-trigger:hover {{ transform: translateY(-2px) scale(1.03); box-shadow: none; }}
    .bot-trigger img {{ width: 100%; height: 100%; object-fit: contain; }}
    .bot-bubble {{
      position: fixed;
      right: 28px;
      bottom: 98px;
      z-index: 2000;
      max-width: 230px;
      padding: 11px 13px;
      color: var(--text);
      background: #ffffff;
      border: 1px solid var(--border);
      border-radius: 16px 16px 4px 16px;
      box-shadow: var(--shadow-md);
      font-size: 12px;
      font-weight: 700;
      opacity: 0;
      pointer-events: none;
      transform: translateY(12px);
      transition: all 220ms ease;
    }}
    .bot-bubble.show {{ opacity: 1; transform: translateY(0); }}
    .bot-window {{
      position: fixed;
      right: 24px;
      bottom: 98px;
      z-index: 1999;
      display: flex;
      flex-direction: column;
      width: min(380px, calc(100vw - 32px));
      height: min(550px, calc(100vh - 130px));
      overflow: hidden;
      background: #ffffff;
      border: 1px solid var(--border);
      border-radius: 22px;
      box-shadow: var(--shadow-xl);
      opacity: 0;
      visibility: hidden;
      transform: scale(0.96) translateY(16px);
      transition: all 220ms ease;
    }}
    .bot-window.open {{ opacity: 1; visibility: visible; transform: scale(1) translateY(0); }}
    .bot-header {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      padding: 16px;
      color: var(--text);
      background: #f8fafc;
      border-bottom: 1px solid var(--border);
    }}
    .bot-header h3 {{
      margin: 0;
      font-size: 11px;
      font-weight: 900;
      letter-spacing: 0.13em;
      text-transform: uppercase;
    }}
    .bot-messages {{
      flex: 1;
      display: flex;
      flex-direction: column;
      gap: 10px;
      overflow-y: auto;
      padding: 16px;
      background: #ffffff;
    }}
    .msg {{
      max-width: 84%;
      padding: 10px 12px;
      border-radius: 16px;
      font-size: 13px;
      font-weight: 600;
      line-height: 1.45;
    }}
    .msg.bot {{
      align-self: flex-start;
      color: var(--text);
      background: #f8fafc;
      border: 1px solid var(--border);
      border-bottom-left-radius: 5px;
    }}
    .msg.user {{
      align-self: flex-end;
      color: #ffffff;
      background: var(--slate);
      border-bottom-right-radius: 5px;
    }}
    .bot-input-area {{
      display: flex;
      gap: 8px;
      padding: 12px;
      background: #ffffff;
      border-top: 1px solid var(--border);
    }}
    .send-btn {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 40px;
      min-width: 40px;
      height: 40px;
      color: #ffffff;
      background: var(--slate);
      border: 1px solid var(--slate);
      border-radius: 12px;
      font-size: 16px;
    }}
    @media (max-width: 980px) {{
      .upload-zone {{ width: fit-content; }}
      .action-group {{ width: auto; }}
    }}
    @media (max-width: 640px) {{
      .shell {{ padding: 12px; }}
      .hero {{ align-items: flex-start; flex-wrap: wrap; border-radius: 20px; }}
      .myntra-btn {{ width: 100%; margin-left: 0; }}
      .global-inputs {{ width: 460px; max-width: 100%; }}
      .margin-input-group {{ width: 210px; max-width: calc(100vw - 44px); }}
      .tabs-container {{ flex-direction: column; }}
      .single-results-grid {{ grid-template-columns: 1fr; }}
      .popup-actions {{ flex-direction: column-reverse; }}
      .popup-actions .primary-btn {{ width: 100%; }}
    }}
    body, body * {{
      font-style: normal !important;
      font-weight: 400 !important;
      letter-spacing: 0 !important;
    }}
    body .primary-btn,
    body .tab-btn,
    body .toggle-label,
    body #pagination-controls label,
    body #pagination-controls span,
    body .file-label .text,
    body .sample-link,
    body .result-card h4,
    body th,
    body td,
    body .page-btn,
    body .popup-card h3,
    body .bot-header h3,
    body .msg,
    body .myntra-btn {{
      font-style: normal !important;
      font-weight: 400 !important;
      letter-spacing: 0 !important;
    }}
    body .margin-input-group input,
    body #export-file-name,
    body .bot-input-area input,
    body select {{
      font-style: normal !important;
      font-weight: 400 !important;
      letter-spacing: 0 !important;
    }}
    body #single-container .margin-input-group label {{
      font-style: normal !important;
      font-weight: 400 !important;
      letter-spacing: 0 !important;
    }}
    body,
    body h1,
    body input,
    body select,
    body textarea,
    body .tab-btn:not(.active),
    body .margin-input-group label,
    body .toggle-label,
    body #pagination-controls label,
    body #pagination-controls span,
    body .file-label,
    body .file-label .text,
    body .sample-link,
    body .result-card h4,
    body .result-card p,
    body th,
    body td,
    body .empty,
    body .popup-card p,
    body .bot-bubble,
    body .msg.bot {{
      color: #000814 !important;
    }}
    body input::placeholder,
    body textarea::placeholder {{
      color: #475569 !important;
      opacity: 1;
    }}
    h1 {{
      font-size: clamp(26px, 3vw, 42px) !important;
      line-height: 1.1 !important;
    }}
    .result-card p {{
      font-size: 24px !important;
      line-height: 1.1 !important;
    }}
  </style>
  <script src="https://cdn.jsdelivr.net/npm/exceljs@4.4.0/dist/exceljs.min.js"></script>
</head>
<body>
  <main class="shell">
    <section class="hero">
      <img src="/logo.png" class="brand-logo" alt="AJIO Logo">
      <h1>AJIO Cost calculator</h1>
      <a href="https://myntra-cost-claculator.onrender.com" target="_blank" class="myntra-btn">
        <svg viewBox="0 0 24 24" width="20" height="20" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
          <rect x="4" y="2" width="16" height="20" rx="2" ry="2"></rect>
          <line x1="8" y1="6" x2="16" y2="6"></line>
          <line x1="16" y1="10" x2="16" y2="10"></line>
          <line x1="12" y1="10" x2="12" y2="10"></line>
          <line x1="8" y1="10" x2="8" y2="10"></line>
          <line x1="16" y1="14" x2="16" y2="14"></line>
          <line x1="12" y1="14" x2="12" y2="14"></line>
          <line x1="8" y1="14" x2="8" y2="14"></line>
          <line x1="16" y1="18" x2="16" y2="18"></line>
          <line x1="12" y1="18" x2="12" y2="18"></line>
          <line x1="8" y1="18" x2="8" y2="18"></line>
        </svg>
        MYNTRA
      </a>
    </section>
    <div class="global-inputs">
      <div class="margin-input-group">
        <label for="company_margin">COMPANY PROFIT MARGIN %</label>
        <input type="number" id="company_margin" name="company_margin" placeholder="0" step="0.01">
      </div>
      <div class="margin-input-group">
        <label for="gst_margin">SALE GST %</label>
        <input type="number" id="gst_margin" name="gst_margin" placeholder="0" step="0.01">
      </div>
      <div class="margin-input-group">
        <label for="discount_margin">SALE DISCOUNT AMT%</label>
        <input type="number" id="discount_margin" name="discount_margin" placeholder="0" step="0.01">
      </div>
      <div class="margin-input-group">
        <label for="ajio_margin">AJIO MARGIN %</label>
        <input type="number" id="ajio_margin" name="ajio_margin" placeholder="0" step="0.01">
      </div>
    </div>

    <div class="tabs-container">
      <button class="tab-btn" id="tab-single" style="display: flex; align-items: center; justify-content: center; gap: 6px;">
        <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M8 12.5V4a2 2 0 0 1 4 0v6.5"></path><path d="M12 10.5a2 2 0 0 1 4 0v3.5"></path><path d="M16 12.5a2 2 0 0 1 4 0v4.5a6 6 0 0 1-6 6H9.5a6 6 0 0 1-5.4-3.4l-2.6-5.2a2 2 0 0 1 3.6-1.8l2.9 5.8"></path></svg> Single Item Calculator
      </button>
      <button class="tab-btn" id="tab-bulk" style="display: flex; align-items: center; justify-content: center; gap: 6px;">
        <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="8.5" y="2" width="7" height="8"></rect><path d="M10.5 2v3l1.5-1.5L13.5 5V2"></path><rect x="4" y="10" width="7" height="8"></rect><path d="M6 10v3l1.5-1.5L9 13v-3"></path><rect x="13" y="10" width="7" height="8"></rect><path d="M15 10v3l1.5-1.5L18 13v-3"></path><path d="M2 20h20"></path><path d="M5 20v2"></path><path d="M12 20v2"></path><path d="M19 20v2"></path><path d="M10 7h4M5.5 15h4M14.5 15h4"></path></svg> Bulk Excel Processor
      </button>
    </div>

    <div id="single-container" style="display: none;">
      <div style="display: flex; flex-wrap: wrap; gap: 10px; padding: 0 34px 10px; align-items: center;">
        <div class="margin-input-group" style="background: var(--surface);">
          <label for="single_y_val">FINAL PURCHASE COST (Y)</label>
          <input type="number" id="single_y_val" placeholder="0" step="0.01">
        </div>
        <div class="margin-input-group" style="background: var(--surface);">
          <label for="single_purchase_gst">PURCHASE GST %</label>
          <input type="number" id="single_purchase_gst" placeholder="0" step="0.01">
        </div>
        <div class="action-group" style="margin-left: auto; display: flex; gap: 10px;">
          <button onclick="exportSingleItemExcel()" class="primary-btn btn-export" style="display: flex; align-items: center; gap: 6px;">
            <svg viewBox="0 0 24 24" width="18" height="18" fill="currentColor"><path d="M2.854 2.147A1 1 0 0 0 2 3v18a1 1 0 0 0 .854.853l8 1.5A1 1 0 0 0 12 22.34V1.66a1 1 0 0 0-1.146-.987l-8 1.474ZM10 4.22v15.56l-6-1.125V5.345l6-1.125ZM22 4h-9v2h9V4Zm0 4h-9v2h9V8Zm0 4h-9v2h9v-2Zm0 4h-9v2h9v-2Zm-7-2h2v-2h-2v2Zm0-4h2V8h-2v2Zm0-4h2V4h-2v2Z"/><path d="m6.664 16 1.488-3.5h.033L9.67 16h1.25l-2.22-4.875 2.13-4.662H9.55L8.258 9.775h-.034L6.874 6.463H5.66l2.08 4.54L5.438 16h1.226Z"/></svg> Export
          </button>
          <button onclick="clearSingleData()" class="primary-btn btn-clear" style="display: flex; align-items: center; gap: 6px;">
            <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"></polyline><path d="M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"></path><line x1="10" y1="11" x2="10" y2="17"></line><line x1="14" y1="11" x2="14" y2="17"></line></svg> Clear Data
          </button>
        </div>
      </div>
      
      <div class="single-results-grid">
          <div class="result-card"><h4>Purchase GST Amt</h4><p id="res_purchase_gst">0.00</p></div>
          <div class="result-card"><h4>Purchase TP W/O Tax</h4><p id="res_purchase_tp">0.00</p></div>
          <div class="result-card"><h4>COMPANY PROFIT MARGIN % (AA)</h4><p id="res_AA">0.00</p></div>
          <div class="result-card"><h4>SALE TP WITH TAX (AD)</h4><p id="res_AD">0.00</p></div>
          <div class="result-card highlight-card"><h4>MRP (AE)</h4><p id="res_AE">0.00</p></div>
          <div class="result-card"><h4>SALE DISCOUNT AMT (AF)</h4><p id="res_AF">0.00</p></div>
          <div class="result-card"><h4>ASP (GROSS) (AG)</h4><p id="res_AG">0.00</p></div>
          <div class="result-card"><h4>GST% (AH)</h4><p id="res_AH_pct">0%</p></div>
          <div class="result-card"><h4>GST amount (AI)</h4><p id="res_AI">0.00</p></div>
          <div class="result-card"><h4>NET SALE (AJ)</h4><p id="res_AJ">0.00</p></div>
          <div class="result-card"><h4>AJIO MARGIN 34% (AK)</h4><p id="res_AK">0.00</p></div>
          <div class="result-card"><h4>PURCHASE (AL)</h4><p id="res_AL">0.00</p></div>
          <div class="result-card"><h4>GST%2 (AM)</h4><p id="res_AM_pct">0%</p></div>
          <div class="result-card"><h4>GST2 amount (AN)</h4><p id="res_AN">0.00</p></div>
          <div class="result-card highlight-card"><h4>BANK SETTLEMENT (AO)</h4><p id="res_AO">0.00</p></div>
      </div>
    </div>

    <div id="bulk-container" style="display: none;">
      <div class="upload-zone">
        <form method="post" enctype="multipart/form-data" id="bulk-upload-form">
          <div class="file-input-wrapper" style="max-width: 250px;">
            <input type="file" id="bulk-file-input" name="file" accept=".xlsx, .xls" required>
            <label for="bulk-file-input" class="file-label" style="padding: 8px 15px;">
              <span class="icon">📁</span>
              <span class="text" id="file-name-display">Choose File</span>
            </label>
          </div>
          
          <label class="toggle-container">
            <span class="toggle-label">Purchase GST</span>
            <div class="switch">
              <input type="checkbox" name="purchase_gst_toggle">
              <span class="slider"></span>
            </div>
          </label>
          
          <button type="submit" class="primary-btn" style="display: flex; align-items: center; gap: 6px;">
            <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="17 8 12 3 7 8"></polyline><line x1="12" y1="3" x2="12" y2="15"></line></svg> Upload
          </button>
          <a href="/sample-excel" class="sample-link">Download Sample Excel</a>
        </form>

        <div class="action-group">
          <button onclick="exportToExcel()" class="primary-btn btn-export" style="display: flex; align-items: center; gap: 6px;">
            <svg viewBox="0 0 24 24" width="18" height="18" fill="currentColor"><path d="M2.854 2.147A1 1 0 0 0 2 3v18a1 1 0 0 0 .854.853l8 1.5A1 1 0 0 0 12 22.34V1.66a1 1 0 0 0-1.146-.987l-8 1.474ZM10 4.22v15.56l-6-1.125V5.345l6-1.125ZM22 4h-9v2h9V4Zm0 4h-9v2h9V8Zm0 4h-9v2h9v-2Zm0 4h-9v2h9v-2Zm-7-2h2v-2h-2v2Zm0-4h2V8h-2v2Zm0-4h2V4h-2v2Z"/><path d="m6.664 16 1.488-3.5h.033L9.67 16h1.25l-2.22-4.875 2.13-4.662H9.55L8.258 9.775h-.034L6.874 6.463H5.66l2.08 4.54L5.438 16h1.226Z"/></svg> Export
          </button>
          <button onclick="openClearPopup()" class="primary-btn btn-clear" style="display: flex; align-items: center; gap: 6px;">
            <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"></polyline><path d="M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"></path><line x1="10" y1="11" x2="10" y2="17"></line><line x1="14" y1="11" x2="14" y2="17"></line></svg> Clear Data
          </button>
        </div>
      </div>
      
      {f'<div style="padding: 0 40px;"><p class="file-info">{html.escape(message)}</p></div>' if message else ''}
      
      <div style="padding: 0;">
        <div class="table-wrap">
          {table_html}
        </div>
        
        <div style="padding: 20px 0; display: flex; justify-content: center; margin-bottom: 40px;">
          <div id="pagination-controls" style="display: none; align-items: center; gap: 20px; background: rgba(255,255,255,0.5); padding: 10px 25px; border-radius: 20px; border: 1px solid rgba(255,255,255,0.8); backdrop-filter: blur(10px);">
              <div style="display:flex; align-items:center; gap:8px; margin-right:8px;">
                <label for="page-size-select" style="font-weight:700; color:#4762b4;">Show</label>
                <select id="page-size-select" style="padding:6px 8px; border-radius:8px; border:1px solid rgba(0,0,0,0.06); background:white; font-weight:700; color:#4762b4;">
                  <option value="10">10</option>
                  <option value="25">25</option>
                  <option value="50" selected>50</option>
                  <option value="100">100</option>
                  <option value="all">All</option>
                </select>
                <span style="font-weight:700; color:#4762b4;">entries</span>
              </div>
              <button id="prev-btn" class="page-btn" type="button" style="padding: 8px 20px; border-radius: 12px; font-weight: 700;">Prev</button>
              <span id="page-info" style="font-size: 15px; font-weight: 800; color: #4762b4;"></span>
              <button id="next-btn" class="page-btn" type="button" style="padding: 8px 20px; border-radius: 12px; font-weight: 700;">Next</button>
            </div>
        </div>
      </div>
    </div>
  </main>
  <div class="bot-bubble" id="bot-bubble">Hii I'am AJ your friend how can i help you</div>
  <div class="bot-trigger" onclick="toggleChat()">
    <img src="/chatbot.png" alt="AJ">
  </div>
  <div class="bot-window" id="bot-window">
    <div class="bot-header">
      <h3>AJ - Dashboard Friend</h3>
      <span style="cursor:pointer; font-size: 20px;" onclick="toggleChat()">&times;</span>
    </div>
    <div class="bot-messages" id="bot-messages">
      <div class="msg bot">Hello! I'm AJ, your dashboard friend. ✨ Before we start, may I know your name? 😊</div>
    </div>
    <div class="bot-input-area">
      <input type="text" id="bot-input" placeholder="Type your message..." onkeydown="if(event.key==='Enter') sendMessage()">
      <button class="send-btn" onclick="sendMessage()">➤</button>
    </div>
  </div>

  <div class="popup-backdrop" id="notice-popup">
    <div class="popup-card">
      <h3 id="notice-title">Notice</h3>
      <p id="notice-message"></p>
      <div class="popup-actions">
        <button type="button" id="notice-close" class="primary-btn btn-back">Close</button>
      </div>
    </div>
  </div>

  <div class="popup-backdrop" id="export-popup">
    <div class="popup-card">
      <h3>Export Excel</h3>
      <p>Enter the name for the exported file. File will be downloaded in .xlsx format.</p>
      <input type="text" id="export-file-name" placeholder="ajio_bulk_export">
      <div class="popup-actions">
        <button type="button" id="cancel-export" class="primary-btn btn-back">Cancel</button>
        <button type="button" id="confirm-export" class="primary-btn btn-export">Download File</button>
      </div>
    </div>
  </div>
  <div class="popup-backdrop" id="clear-popup">
    <div class="popup-card">
      <h3>Clear Data</h3>
      <p>Do you want to delete the existing data and upload new data?</p>
      <div class="popup-actions">
        <button type="button" id="cancel-clear" class="primary-btn btn-back">Cancel</button>
        <button type="button" id="confirm-clear" class="primary-btn btn-clear">Yes, Clear Data</button>
      </div>
    </div>
  </div>
  <script>
    document.addEventListener('DOMContentLoaded', () => {{
      const marginInput = document.getElementById('company_margin');
      const gstInput = document.getElementById('gst_margin');
      const discountInput = document.getElementById('discount_margin');
      const ajioInput = document.getElementById('ajio_margin');
      const tableBody = document.getElementById('table-body');
      
      const fileInput = document.getElementById('bulk-file-input');
      const fileNameDisplay = document.getElementById('file-name-display');
      const uploadForm = document.getElementById('bulk-upload-form');

      if (fileInput && fileNameDisplay) {{
        fileInput.addEventListener('change', (e) => {{
          const fileName = e.target.files[0]?.name || 'Choose File';
          fileNameDisplay.textContent = fileName;
        }});
      }}

      if (uploadForm) {{
        uploadForm.addEventListener('submit', () => {{
          const btn = uploadForm.querySelector('button[type="submit"]');
          if (btn) {{
            btn.disabled = true;
            btn.textContent = 'Uploading...';
          }}
        }});
      }}
      const prevBtn = document.getElementById('prev-btn');
      const nextBtn = document.getElementById('next-btn');
      const pageInfo = document.getElementById('page-info');
      const paginationControls = document.getElementById('pagination-controls');
      const exportPopup = document.getElementById('export-popup');
      const openExportPopupBtn = document.getElementById('open-export-popup');
      const clearDataBtn = document.getElementById('clear-data-btn');
      const cancelExportBtn = document.getElementById('cancel-export');
      const confirmExportBtn = document.getElementById('confirm-export');
      const exportFileNameInput = document.getElementById('export-file-name');
      
      const clearPopup = document.getElementById('clear-popup');
      const cancelClearBtn = document.getElementById('cancel-clear');
      const confirmClearBtn = document.getElementById('confirm-clear');
      const noticePopup = document.getElementById('notice-popup');
      const noticeTitle = document.getElementById('notice-title');
      const noticeMessage = document.getElementById('notice-message');
      const noticeClose = document.getElementById('notice-close');

      function showNotice(message, title = 'Notice') {{
        if (!noticePopup || !noticeTitle || !noticeMessage) return;
        noticeTitle.textContent = title;
        noticeMessage.textContent = message;
        noticePopup.classList.add('open');
      }}

      if (noticeClose) {{
        noticeClose.addEventListener('click', () => noticePopup.classList.remove('open'));
      }}
      if (noticePopup) {{
        noticePopup.addEventListener('click', (event) => {{
          if (event.target === noticePopup) noticePopup.classList.remove('open');
        }});
      }}
      
      let currentPage = 1;
      let pageSize = 50;
      const tableData = typeof globalTableData !== 'undefined' ? globalTableData : [];
      const exportHeaders = typeof globalHeaders !== 'undefined' ? globalHeaders : [];
      const hasUploadedData = tableData.length > 0;

      // Initialize pageSize from localStorage (supports 'all')
      const savedPageSize = localStorage.getItem('ajio_page_size');
      if (savedPageSize === 'all') {{
        pageSize = tableData.length || 50;
      }} else if (savedPageSize) {{
        const parsed = parseInt(savedPageSize, 10);
        if (!isNaN(parsed) && parsed > 0) pageSize = parsed;
      }}

      if (tableData.length > 0) {{
        paginationControls.style.display = 'flex';
        const pageSizeSelect = document.getElementById('page-size-select');
        if (pageSizeSelect) {{
          // Set select to match saved value (or default)
          if (savedPageSize === 'all') {{
            pageSizeSelect.value = 'all';
          }} else {{
            pageSizeSelect.value = String(pageSize);
          }}

          pageSizeSelect.addEventListener('change', (e) => {{
            const v = e.target.value;
            if (v === 'all') {{
              pageSize = tableData.length || 1;
            }} else {{
              pageSize = parseInt(v, 10) || 50;
            }}
            localStorage.setItem('ajio_page_size', v);
            currentPage = 1;
            renderPage();
          }});
        }}
      }}

      function escapeHtml(text) {{
        const map = {{
          '&': '&amp;',
          '<': '&lt;',
          '>': '&gt;',
          '"': '&quot;',
          "'": '&#039;'
        }};
        return String(text).replace(/[&<>"']/g, function(m) {{ return map[m]; }});
      }}

      function sanitizeFilename(name) {{
        const cleaned = String(name || '')
          .replace(/[<>:"/\\\\|?*]+/g, '')
          .trim()
          .replace(/\\.+$/, '');
        return cleaned || 'ajio_bulk_export';
      }}

      function calculateSettlementFromMRP(mrp, discount, ajio) {{
        const discountAmt = Math.round((mrp * discount) / 100);
        const aspGross = mrp - discountAmt;
        const gstPct = aspGross > 2499 ? 0.18 : 0.05;
        const gstAmt = aspGross - (aspGross / (1 + gstPct));
        const ajioMarginAmt = Math.round(((aspGross * ajio) / 100) * 100) / 100;
        
        const purchase = aspGross - gstAmt - ajioMarginAmt;
        const gst2Pct = purchase > 2499 ? 0.18 : 0.05;
        const gst2Amt = purchase * gst2Pct;
        return purchase + gst2Amt;
      }}

      function calculateAE_raw(targetAD, discount, ajio) {{
        if (targetAD <= 0) return 0;
        const d = (typeof discount === 'number' && !isNaN(discount)) ? discount : (parseFloat(discountInput ? discountInput.value.trim() : '') || 65);
        const m = (typeof ajio === 'number' && !isNaN(ajio)) ? ajio : (parseFloat(ajioInput ? ajioInput.value.trim() : '') || 34);

        let low = 0;
        let high = targetAD * 30;
        for (let i = 0; i < 60; i++) {{
          const mid = (low + high) / 2;
          if (calculateSettlementFromMRP(mid, d, m) < targetAD) {{
            low = mid;
          }} else {{
            high = mid;
          }}
        }}
        return (low + high) / 2;
      }}

      const tabSingle = document.getElementById('tab-single');
      const tabBulk = document.getElementById('tab-bulk');
      const containerSingle = document.getElementById('single-container');
      const containerBulk = document.getElementById('bulk-container');
      function activateTab(mode) {{
        if (!tabSingle || !tabBulk) return;
        const showBulk = mode === 'bulk';
        tabBulk.classList.toggle('active', showBulk);
        tabSingle.classList.toggle('active', !showBulk);
        containerBulk.style.display = showBulk ? 'block' : 'none';
        containerSingle.style.display = showBulk ? 'none' : 'block';
        localStorage.setItem('ajio_active_tab', mode);
      }}

      if (tabSingle && tabBulk) {{
        tabSingle.addEventListener('click', () => {{
          activateTab('single');
        }});

        tabBulk.addEventListener('click', () => {{
          activateTab('bulk');
        }});
      }}

      const savedTab = localStorage.getItem('ajio_active_tab');
      const serverDefaultTab = "{default_tab}";
      let initialTab = 'single';
      
      if (serverDefaultTab === 'bulk') {{
        initialTab = 'bulk';
      }} else if (hasUploadedData) {{
        initialTab = 'bulk';
      }} else if (savedTab) {{
        initialTab = savedTab;
      }}
      
      activateTab(initialTab);

      const singleYInput = document.getElementById('single_y_val');
      const singlePurchGstInput = document.getElementById('single_purchase_gst');

      function updateSingleMode() {{
        if (!singleYInput) return;
        
        const yValue = parseFloat(singleYInput.value.trim()) || 0;
        const purchGst = parseFloat(singlePurchGstInput.value.trim()) || 0;
        const margin = parseFloat(marginInput.value.trim()) || 0;
        const discount = parseFloat(discountInput.value.trim()) || 0;
        const ajio = parseFloat(ajioInput.value.trim()) || 0;
        
        // Purchase Side
        const purchGstAmt = Math.round(((yValue * purchGst) / (100 + purchGst)) * 100) / 100;
        const purchTp = Math.round((yValue - purchGstAmt) * 100) / 100;
        
        // Sales Side (Target calculations)
        const valAA = Math.round(((yValue * margin) / 100) * 100) / 100;
        const valAD = Math.round((yValue + valAA) * 100) / 100;
        
        // GoalSeek solver for MRP
        const valAE_raw = calculateAE_raw(valAD, discount, ajio);
        
        // Final calculations based on RAW MRP for precision
        const valAF_raw = (valAE_raw * discount) / 100;
        const valAG_raw = valAE_raw - valAF_raw;
        const valAH_num = valAG_raw > 2499 ? 0.18 : 0.05;
        const valAH_pct = (valAH_num * 100) + '%';
        const valAI_raw = valAG_raw - (valAG_raw / (1 + valAH_num));
        const valAJ_raw = valAG_raw - valAI_raw;
        const valAK_raw = (valAG_raw * ajio) / 100;
        const valAL_raw = valAG_raw - valAI_raw - valAK_raw;
        const valAM_num = valAL_raw > 2499 ? 0.18 : 0.05;
        const valAM_pct = (valAM_num * 100) + '%';
        const valAN_raw = valAL_raw * valAM_num;
        const valAO_raw = valAL_raw + valAN_raw;

        document.getElementById('res_purchase_gst').innerText = purchGstAmt.toFixed(2);
        document.getElementById('res_purchase_tp').innerText = purchTp.toFixed(2);
        document.getElementById('res_AA').innerText = valAA.toFixed(2);
        document.getElementById('res_AD').innerText = valAD.toFixed(2);
        
        // Display Rounded Values as requested
        document.getElementById('res_AE').innerText = Math.round(valAE_raw).toFixed(0);
        document.getElementById('res_AF').innerText = Math.round(valAF_raw).toFixed(0);
        document.getElementById('res_AG').innerText = Math.round(valAG_raw).toFixed(0);
        
        document.getElementById('res_AH_pct').innerText = valAH_pct;
        document.getElementById('res_AI').innerText = valAI_raw.toFixed(2);
        document.getElementById('res_AJ').innerText = valAJ_raw.toFixed(2);
        document.getElementById('res_AK').innerText = valAK_raw.toFixed(2);
        document.getElementById('res_AL').innerText = valAL_raw.toFixed(2);
        document.getElementById('res_AM_pct').innerText = valAM_pct;
        document.getElementById('res_AN').innerText = valAN_raw.toFixed(2);
        document.getElementById('res_AO').innerText = valAO_raw.toFixed(2);
      }}

      if (singleYInput) singleYInput.addEventListener('input', updateSingleMode);
      if (singlePurchGstInput) singlePurchGstInput.addEventListener('input', updateSingleMode);

      window.clearSingleData = () => {{
        if (singleYInput) singleYInput.value = '';
        if (singlePurchGstInput) singlePurchGstInput.value = '';
        updateSingleMode();
      }};

      window.exportSingleItemExcel = async () => {{
        if (typeof ExcelJS === 'undefined') {{
          showNotice('Excel library not loaded.', 'Export unavailable');
          return;
        }}
        
        const workbook = new ExcelJS.Workbook();
        workbook.creator = 'AJIO Cost Calculator';
        workbook.created = new Date();

        const ws = workbook.addWorksheet('Single Item Dashboard', {{
          views: [{{ showGridLines: false }}]
        }});

        ws.columns = [
          {{ width: 18 }}, {{ width: 14 }}, {{ width: 3 }},
          {{ width: 18 }}, {{ width: 14 }}, {{ width: 3 }},
          {{ width: 18 }}, {{ width: 14 }}, {{ width: 3 }},
          {{ width: 18 }}, {{ width: 14 }}
        ];

        const palettes = {{
          yellow: {{ fill: 'FFFEFCE8', border: 'FFFDE68A', accent: 'FF854D0E' }},
          purple: {{ fill: 'FFFAF5FF', border: 'FFE9D5FF', accent: 'FF6B21A8' }},
          rose: {{ fill: 'FFFFF1F2', border: 'FFFECDD3', accent: 'FF9F1239' }},
          blue: {{ fill: 'FFEFF6FF', border: 'FFBFDBFE', accent: 'FF1E40AF' }},
          green: {{ fill: 'FFECFDF5', border: 'FFA7F3D0', accent: 'FF065F46' }},
          orange: {{ fill: 'FFFFF7ED', border: 'FFFED7AA', accent: 'FF9A3412' }},
          indigo: {{ fill: 'FFEEF2FF', border: 'FFC7D2FE', accent: 'FF3730A3' }}
        }};

        const dark = 'FF000814';
        const lightBorder = 'FFE2E8F0';
        const sectionFill = 'FFF8FAFC';

        function applyBox(row, col, label, value, theme, numFmt) {{
          const palette = palettes[theme] || palettes.blue;
          ws.mergeCells(row, col, row, col + 1);
          ws.mergeCells(row + 1, col, row + 2, col + 1);

          for (let r = row; r <= row + 2; r++) {{
            for (let c = col; c <= col + 1; c++) {{
              const cell = ws.getCell(r, c);
              cell.fill = {{ type: 'pattern', pattern: 'solid', fgColor: {{ argb: palette.fill }} }};
              cell.border = {{
                top: {{ style: 'thin', color: {{ argb: palette.border }} }},
                left: {{ style: 'thin', color: {{ argb: palette.border }} }},
                bottom: {{ style: 'thin', color: {{ argb: palette.border }} }},
                right: {{ style: 'thin', color: {{ argb: palette.border }} }}
              }};
            }}
          }}

          const labelCell = ws.getCell(row, col);
          labelCell.value = label;
          labelCell.font = {{ name: 'Calibri', size: 9, bold: false, color: {{ argb: palette.accent }} }};
          labelCell.alignment = {{ vertical: 'middle', horizontal: 'left', wrapText: true }};

          const valueCell = ws.getCell(row + 1, col);
          valueCell.value = value;
          if (numFmt) valueCell.numFmt = numFmt;
          valueCell.font = {{ name: 'Calibri', size: 18, bold: false, color: {{ argb: dark }} }};
          valueCell.alignment = {{ vertical: 'middle', horizontal: 'right' }};
        }}

        function sectionTitle(row, title) {{
          ws.mergeCells(row, 1, row, 11);
          const cell = ws.getCell(row, 1);
          cell.value = title;
          cell.font = {{ name: 'Calibri', size: 12, bold: false, color: {{ argb: dark }} }};
          cell.fill = {{ type: 'pattern', pattern: 'solid', fgColor: {{ argb: sectionFill }} }};
          cell.border = {{
            top: {{ style: 'thin', color: {{ argb: lightBorder }} }},
            left: {{ style: 'thin', color: {{ argb: lightBorder }} }},
            bottom: {{ style: 'thin', color: {{ argb: lightBorder }} }},
            right: {{ style: 'thin', color: {{ argb: lightBorder }} }}
          }};
          cell.alignment = {{ vertical: 'middle', horizontal: 'left' }};
        }}

        ws.mergeCells(1, 2, 1, 11);
        ws.getCell(1, 2).value = 'AJIO Cost Calculator';
        ws.getCell(1, 2).font = {{ name: 'Calibri', size: 22, bold: false, color: {{ argb: dark }} }};
        ws.getCell(1, 2).alignment = {{ vertical: 'middle', horizontal: 'left' }};
        ws.getRow(1).height = 42;

        try {{
          const logoResponse = await fetch('/logo.png');
          const logoBuffer = await logoResponse.arrayBuffer();
          const logoId = workbook.addImage({{ buffer: logoBuffer, extension: 'png' }});
          ws.addImage(logoId, {{
            tl: {{ col: 0.15, row: 0.12 }},
            ext: {{ width: 38, height: 38 }}
          }});
        }} catch (err) {{
          console.warn('Logo could not be added to export:', err);
        }}

        sectionTitle(3, 'Input Summary');
        const singleY = parseFloat(document.getElementById('single_y_val')?.value) || 0;
        const singlePurchGst = parseFloat(document.getElementById('single_purchase_gst')?.value) || 0;
        const compMargin = parseFloat(document.getElementById('company_margin')?.value) || 0;
        const saleGst = parseFloat(document.getElementById('gst_margin')?.value) || 0;
        const saleDiscount = parseFloat(document.getElementById('discount_margin')?.value) || 0;
        const ajioMargin = parseFloat(document.getElementById('ajio_margin')?.value) || 0;

        const inputCards = [
          {{ label: 'FINAL PURCHASE COST (Y)', value: singleY, theme: 'yellow', numFmt: '#,##0.00' }},
          {{ label: 'PURCHASE GST %', value: singlePurchGst, theme: 'purple', numFmt: '0.00' }},
          {{ label: 'COMPANY PROFIT MARGIN %', value: compMargin, theme: 'rose', numFmt: '0.00' }},
          {{ label: 'SALE GST %', value: saleGst, theme: 'blue', numFmt: '0.00' }},
          {{ label: 'SALE DISCOUNT AMT%', value: saleDiscount, theme: 'green', numFmt: '0.00' }},
          {{ label: 'AJIO MARGIN %', value: ajioMargin, theme: 'orange', numFmt: '0.00' }}
        ];

        const cardColumns = [1, 4, 7, 10];
        inputCards.forEach((card, index) => {{
          const row = index < 4 ? 5 : 9;
          const col = cardColumns[index % 4];
          applyBox(row, col, card.label, card.value, card.theme, card.numFmt);
        }});

        sectionTitle(13, 'Calculated Results');
        const resPurchGst = parseFloat(document.getElementById('res_purchase_gst')?.innerText) || 0;
        const resPurchTp = parseFloat(document.getElementById('res_purchase_tp')?.innerText) || 0;
        const resAA = parseFloat(document.getElementById('res_AA')?.innerText) || 0;
        const resAD = parseFloat(document.getElementById('res_AD')?.innerText) || 0;
        const resAE = parseFloat(document.getElementById('res_AE')?.innerText) || 0;
        const resAF = parseFloat(document.getElementById('res_AF')?.innerText) || 0;
        const resAG = parseFloat(document.getElementById('res_AG')?.innerText) || 0;
        const resAHText = document.getElementById('res_AH_pct')?.innerText || '5%';
        const resAHNum = resAHText.includes('18') ? 0.18 : 0.05;
        const resAI = parseFloat(document.getElementById('res_AI')?.innerText) || 0;
        const resAJ = parseFloat(document.getElementById('res_AJ')?.innerText) || 0;
        const resAK = parseFloat(document.getElementById('res_AK')?.innerText) || 0;
        const resAL = parseFloat(document.getElementById('res_AL')?.innerText) || 0;
        const resAMText = document.getElementById('res_AM_pct')?.innerText || '5%';
        const resAMNum = resAMText.includes('18') ? 0.18 : 0.05;
        const resAN = parseFloat(document.getElementById('res_AN')?.innerText) || 0;
        const resAO = parseFloat(document.getElementById('res_AO')?.innerText) || 0;

        const singleD = saleDiscount / 100;
        const singleM = ajioMargin / 100;
        const singleT1 = Math.round((2499 * (1 - singleM * 1.05)) * 100) / 100;
        const singleT2 = 2623.95;
        const singleK1 = (1 - singleD) * (1 - singleM * 1.05);
        const singleK2 = (1 - singleD) * (1 / 1.18 - singleM) * 1.05;
        const singleK3 = (1 - singleD) * (1 - singleM * 1.18);
        const singleAeFormula = (singleK1 > 0 && singleK2 > 0 && singleK3 > 0)
          ? 'IF(J16<=0,0,ROUND(IF(J16<=' + singleT1 + ',J16/' + singleK1.toFixed(8) + ',IF(J16<=' + singleT2 + ',J16/' + singleK2.toFixed(8) + ',J16/' + singleK3.toFixed(8) + ')),0))'
          : 'IF(J16<=0,0,ROUND(J16/0.22505,0))';

        const resultCards = [
          {{ label: 'Purchase GST Amt', value: {{ formula: 'ROUND((A6*D6)/(100+D6),2)', result: resPurchGst }}, theme: 'blue', numFmt: '#,##0.00' }},
          {{ label: 'Purchase TP W/O Tax', value: {{ formula: 'ROUND(A6-A16,2)', result: resPurchTp }}, theme: 'blue', numFmt: '#,##0.00' }},
          {{ label: 'COMPANY PROFIT MARGIN % (AA)', value: {{ formula: 'ROUND((A6*G6)/100,2)', result: resAA }}, theme: 'green', numFmt: '#,##0.00' }},
          {{ label: 'SALE TP WITH TAX (AD)', value: {{ formula: 'ROUND(A6+G16,2)', result: resAD }}, theme: 'orange', numFmt: '#,##0.00' }},
          {{ label: 'MRP (AE)', value: resAE, theme: 'orange', numFmt: '#,##0' }},
          {{ label: 'SALE DISCOUNT AMT (AF)', value: {{ formula: 'ROUND((A20*A10)/100,0)', result: resAF }}, theme: 'rose', numFmt: '#,##0' }},
          {{ label: 'ASP (GROSS) (AG)', value: {{ formula: 'A20-D20', result: resAG }}, theme: 'orange', numFmt: '#,##0' }},
          {{ label: 'GST% (AH)', value: {{ formula: 'IF(G20>2499,0.18,0.05)', result: resAHNum }}, theme: 'indigo', numFmt: '0%' }},
          {{ label: 'GST amount (AI)', value: {{ formula: 'ROUND(G20-(G20/(1+J20)),2)', result: resAI }}, theme: 'indigo', numFmt: '#,##0.00' }},
          {{ label: 'NET SALE (AJ)', value: {{ formula: 'ROUND(G20-A24,2)', result: resAJ }}, theme: 'orange', numFmt: '#,##0.00' }},
          {{ label: 'AJIO MARGIN 34% (AK)', value: {{ formula: 'ROUND((G20*D10)/100,2)', result: resAK }}, theme: 'green', numFmt: '#,##0.00' }},
          {{ label: 'PURCHASE (AL)', value: {{ formula: 'ROUND(G20-A24-G24,2)', result: resAL }}, theme: 'blue', numFmt: '#,##0.00' }},
          {{ label: 'GST%2 (AM)', value: {{ formula: 'IF(J24>2499,0.18,0.05)', result: resAMNum }}, theme: 'indigo', numFmt: '0%' }},
          {{ label: 'GST2 amount (AN)', value: {{ formula: 'ROUND(J24*A28,2)', result: resAN }}, theme: 'indigo', numFmt: '#,##0.00' }},
          {{ label: 'BANK SETTLEMENT (AO)', value: {{ formula: 'ROUND(J24+D28,2)', result: resAO }}, theme: 'green', numFmt: '#,##0.00' }}
        ];

        resultCards.forEach((card, index) => {{
          const row = 15 + Math.floor(index / 4) * 4;
          const col = cardColumns[index % 4];
          applyBox(row, col, card.label, card.value, card.theme, card.numFmt);
        }});

        for (let r = 1; r <= 30; r++) {{
          if (!ws.getRow(r).height) ws.getRow(r).height = 18;
        }}
        [5, 9, 15, 19, 23, 27].forEach((row) => {{
          ws.getRow(row).height = 20;
          ws.getRow(row + 1).height = 20;
          ws.getRow(row + 2).height = 20;
        }});
        
        const buffer = await workbook.xlsx.writeBuffer();
        const blob = new Blob([buffer], {{ type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' }});
        const url = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = url;
        a.download = 'ajio_single_item_export.xlsx';
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);
      }};

      function renderPage() {{
        if (tableData.length === 0) return;
        
        const totalPages = Math.ceil(tableData.length / pageSize);
        if (currentPage < 1) currentPage = 1;
        if (currentPage > totalPages) currentPage = totalPages;
        
        pageInfo.innerText = `Page ${{currentPage}} of ${{totalPages}}`;
        prevBtn.disabled = currentPage === 1;
        nextBtn.disabled = currentPage === totalPages;
        
        const start = (currentPage - 1) * pageSize;
        const end = start + pageSize;
        const pageData = tableData.slice(start, end);
        
        let html = '';
        pageData.forEach(row => {{
          let tr = '<tr>';
          row.forEach(cell => {{
            tr += `<td>${{escapeHtml(cell)}}</td>`;
          }});
          tr += '</tr>';
          html += tr;
        }});
        
        tableBody.innerHTML = html;
        updateCompanyProfit();
      }}

      function updateCompanyProfit() {{
        if (!tableData.length) return;
        
        const marginStr = marginInput ? marginInput.value.trim() : '';
        const margin = parseFloat(marginStr) || 0;
        const isMarginValid = !isNaN(parseFloat(marginStr));

        const gstStr = gstInput ? gstInput.value.trim() : '';
        const gst = parseFloat(gstStr) || 0;
        const isGstValid = !isNaN(parseFloat(gstStr));

        const discountStr = discountInput ? discountInput.value.trim() : '';
        const discount = parseFloat(discountStr) || 0;
        const isDiscountValid = !isNaN(parseFloat(discountStr));

        const ajioStr = ajioInput ? ajioInput.value.trim() : '';
        const ajio = parseFloat(ajioStr) || 0;
        const isAjioValid = !isNaN(parseFloat(ajioStr));

        // Update the entire data array so pagination works correctly
        tableData.forEach((row) => {{
          const yValueStr = row[24];
          const yValue = parseFloat(String(yValueStr).replace(/,/g, '')) || 0;

          const profit = isMarginValid ? Math.round(((yValue * margin) / 100) * 100) / 100 : 0;
          const acValue = yValue + profit;
          const abValue = isGstValid ? Math.round(((acValue * gst) / (100 + gst)) * 100) / 100 : 0;
          const aaValue = acValue - abValue;

          let aeRaw = 0;
          if (acValue > 0) {{
            aeRaw = calculateAE_raw(acValue, discount, ajio);
          }}
          
          const afRaw = (aeRaw * discount) / 100;
          const agRaw = aeRaw - afRaw;
          const ahPctVal = agRaw > 2499 ? 0.18 : 0.05;
          const ahValue = agRaw - (agRaw / (1 + ahPctVal));
          const aiValue = agRaw - ahValue;
          const ajValue = (agRaw * ajio) / 100;
          const akValue = agRaw - ahValue - ajValue;
          const alPctVal = akValue > 2499 ? 0.18 : 0.05;
          const amValue = akValue * alPctVal;
          const anValue = akValue + amValue;

          const anyValid = isMarginValid || isGstValid || aeRaw > 0;
          
          // Store updated values back in the data source
          row[25] = '';
          row[26] = isMarginValid ? profit.toFixed(2) : '';
          row[27] = anyValid ? aaValue.toFixed(2) : '';
          row[28] = isGstValid ? abValue.toFixed(2) : '';
          row[29] = anyValid ? acValue.toFixed(2) : '';
          
          // Display Rounded Values as requested
          row[30] = aeRaw > 0 ? Math.round(aeRaw).toFixed(0) : '';
          row[31] = aeRaw > 0 ? Math.round(afRaw).toFixed(0) : '';
          row[32] = aeRaw > 0 ? Math.round(agRaw).toFixed(0) : '';
          
          row[33] = aeRaw > 0 ? (ahPctVal * 100) + '%' : '';
          row[34] = aeRaw > 0 ? ahValue.toFixed(2) : '';
          row[35] = aeRaw > 0 ? aiValue.toFixed(2) : '';
          row[36] = aeRaw > 0 ? ajValue.toFixed(2) : '';
          row[37] = aeRaw > 0 ? akValue.toFixed(2) : '';
          row[38] = aeRaw > 0 ? (alPctVal * 100) + '%' : '';
          row[39] = aeRaw > 0 ? amValue.toFixed(2) : '';
          row[40] = aeRaw > 0 ? anValue.toFixed(2) : '';
        }});

        // Now sync only the visible table rows in the DOM
        const tableRows = tableBody.querySelectorAll('tr');
        tableRows.forEach((tr, index) => {{
          const dataRowIndex = (currentPage - 1) * pageSize + index;
          if (tableData[dataRowIndex] && tr.cells.length >= 41) {{
            for (let i = 25; i <= 40; i++) {{
              tr.cells[i].innerText = tableData[dataRowIndex][i];
            }}
          }}
        }});
      }}

      function buildExportRows() {{
        return tableData.map((originalRow) => {{
          const row = [...originalRow];
          while (row.length < exportHeaders.length) {{
            row.push('');
          }}

          const yValue = parseFloat(String(row[24] || '').replace(/,/g, '')) || 0;
          const marginStr = marginInput ? marginInput.value.trim() : '';
          const margin = parseFloat(marginStr) || 0;
          const isMarginValid = !isNaN(parseFloat(marginStr));

          const gstStr = gstInput ? gstInput.value.trim() : '';
          const gst = parseFloat(gstStr) || 0;
          const isGstValid = !isNaN(parseFloat(gstStr));

          const discountStr = discountInput ? discountInput.value.trim() : '';
          const discount = parseFloat(discountStr) || 0;
          const isDiscountValid = !isNaN(parseFloat(discountStr));

          const ajioStr = ajioInput ? ajioInput.value.trim() : '';
          const ajio = parseFloat(ajioStr) || 0;
          const isAjioValid = !isNaN(parseFloat(ajioStr));

          const profit = isMarginValid ? Math.round(((yValue * margin) / 100) * 100) / 100 : 0;
          const acValue = yValue + profit;
          const abValue = isGstValid ? Math.round(((acValue * gst) / (100 + gst)) * 100) / 100 : 0;
          const aaValue = acValue - abValue;

          let aeRaw = 0;
          if (acValue > 0) {{
            aeRaw = calculateAE_raw(acValue, discount, ajio);
          }}

          const afRaw = (aeRaw * discount) / 100;
          const agRaw = aeRaw - afRaw;
          const ahPctVal = agRaw > 2499 ? 0.18 : 0.05;
          const ahValue = agRaw - (agRaw / (1 + ahPctVal));
          const aiValue = agRaw - ahValue;
          const ajValue = (agRaw * ajio) / 100;
          const akValue = agRaw - ahValue - ajValue;
          const alPctVal = akValue > 2499 ? 0.18 : 0.05;
          const amValue = akValue * alPctVal;
          const anValue = akValue + amValue;

          row[25] = '';
          row[26] = isMarginValid ? profit.toFixed(2) : '';
          row[27] = (isMarginValid || isGstValid || aeRaw > 0) ? aaValue.toFixed(2) : '';
          row[28] = isGstValid ? abValue.toFixed(2) : '';
          row[29] = (isMarginValid || isGstValid || aeRaw > 0) ? acValue.toFixed(2) : '';
          
          row[30] = aeRaw > 0 ? Math.round(aeRaw).toFixed(0) : '';
          row[31] = aeRaw > 0 ? Math.round(afRaw).toFixed(0) : '';
          row[32] = aeRaw > 0 ? Math.round(agRaw).toFixed(0) : '';
          
          row[33] = aeRaw > 0 ? (ahPctVal * 100) + '%' : '';
          row[34] = aeRaw > 0 ? ahValue.toFixed(2) : '';
          row[35] = aeRaw > 0 ? aiValue.toFixed(2) : '';
          row[36] = aeRaw > 0 ? ajValue.toFixed(2) : '';
          row[37] = aeRaw > 0 ? akValue.toFixed(2) : '';
          row[38] = aeRaw > 0 ? (alPctVal * 100) + '%' : '';
          row[39] = aeRaw > 0 ? amValue.toFixed(2) : '';
          row[40] = aeRaw > 0 ? anValue.toFixed(2) : '';
          return row;
        }});
      }}

      function openExportPopup() {{
        if (!tableData.length) {{
          showNotice('Please upload Excel file first to export data.', 'Upload required');
          return;
        }}
        exportFileNameInput.value = 'ajio_bulk_export';
        exportPopup.classList.add('open');
        exportFileNameInput.focus();
        exportFileNameInput.select();
      }}

      function closeExportPopup() {{
        exportPopup.classList.remove('open');
      }}

      async function exportExcel() {{
        if (!tableData.length) {{
          closeExportPopup();
          return;
        }}

        if (typeof ExcelJS === 'undefined') {{
          showNotice('Excel library load nahi hui. Please page refresh karke try karein.', 'Export unavailable');
          return;
        }}

        const safeName = sanitizeFilename(exportFileNameInput.value);
        const rows = buildExportRows();

        confirmExportBtn.disabled = true;
        confirmExportBtn.textContent = 'Preparing...';

        try {{
          // ── 1. Create workbook & worksheet with freeze top row ──
          const workbook = new ExcelJS.Workbook();
          const ws = workbook.addWorksheet('AJIO Export', {{
            views: [{{ state: 'frozen', ySplit: 1 }}]
          }});

          // ── 2. Add header row ──
          ws.addRow(exportHeaders);

          const marginStr = marginInput ? marginInput.value.trim() : '';
          const margin = parseFloat(marginStr) || 0;
          const isMarginValid = !isNaN(parseFloat(marginStr));

          const gstStr = gstInput ? gstInput.value.trim() : '';
          const gst = parseFloat(gstStr) || 0;
          const isGstValid = !isNaN(parseFloat(gstStr));

          const discountStr = discountInput ? discountInput.value.trim() : '';
          const discount = parseFloat(discountStr) || 0;
          const isDiscountValid = !isNaN(parseFloat(discountStr));

          const ajioStr = ajioInput ? ajioInput.value.trim() : '';
          const ajio = parseFloat(ajioStr) || 0;
          const isAjioValid = !isNaN(parseFloat(ajioStr));

          const D = discount / 100;
          const M = ajio / 100;
          const T1 = Math.round((2499 * (1 - M * 1.05)) * 100) / 100;
          const T2 = 2623.95;
          const K1 = (1 - D) * (1 - M * 1.05);
          const K2 = (1 - D) * (1 / 1.18 - M) * 1.05;
          const K3 = (1 - D) * (1 - M * 1.18);

          // ── 3. Add data rows with dynamic formulas & clean numeric values ──
          rows.forEach((row, rowIndex) => {{
            const R = rowIndex + 2;
            const processedRow = row.map(cell => {{
              if (cell === null || cell === undefined || cell === '') return '';
              const s = String(cell).trim();
              if (s.endsWith('%')) {{
                const num = parseFloat(s);
                if (!isNaN(num)) return num / 100;
              }}
              const cleaned = s.replace(/,/g, '');
              if (cleaned !== '' && !isNaN(Number(cleaned))) return Number(cleaned);
              return s;
            }});
            const addedRow = ws.addRow(processedRow);

            const yRaw = String(row[24] || '').replace(/,/g, '').trim();
            const hasY = yRaw !== '' && !isNaN(Number(yRaw));

            // Purchase side formulas (Tax V = col 22, Purchase Cost W = col 23, Purchase Tax X = col 24, Y = col 25)
            const vRaw = String(row[21] || '').trim();
            if (hasY && vRaw !== '') {{
              const xCell = addedRow.getCell(24);
              const xNum = parseFloat(String(row[23] || '').replace(/,/g, ''));
              xCell.value = {{
                formula: 'IF(OR(Y' + R + '="",V' + R + '=""),"",ROUND(IF(V' + R + '<1,(Y' + R + '*V' + R + ')/(1+V' + R + '),(Y' + R + '*V' + R + ')/(100+V' + R + ')),2))',
                result: isNaN(xNum) ? undefined : xNum
              }};
              xCell.numFmt = '#,##0.00';

              const wCell = addedRow.getCell(23);
              const wNum = parseFloat(String(row[22] || '').replace(/,/g, ''));
              wCell.value = {{
                formula: 'IF(Y' + R + '="","",ROUND(Y' + R + '-X' + R + ',2))',
                result: isNaN(wNum) ? undefined : wNum
              }};
              wCell.numFmt = '#,##0.00';
            }}

            // Sales side formulas (Cols AA=27 to AO=41)
            if (hasY && (isMarginValid || isGstValid || isDiscountValid || isAjioValid)) {{
              const yVal = Number(yRaw);
              const profitNum = isMarginValid ? Math.round(((yVal * margin) / 100) * 100) / 100 : 0;
              const acNum = yVal + profitNum;
              const abNum = isGstValid ? Math.round(((acNum * gst) / (100 + gst)) * 100) / 100 : 0;
              const aaNum = acNum - abNum;

              let aeRaw = 0;
              if (acNum > 0) {{
                aeRaw = calculateAE_raw(acNum, discount, ajio);
              }}
              const afRaw = (aeRaw * discount) / 100;
              const agRaw = aeRaw - afRaw;
              const ahPctVal = agRaw > 2499 ? 0.18 : 0.05;
              const ahValue = agRaw - (agRaw / (1 + ahPctVal));
              const aiValue = agRaw - ahValue;
              const ajValue = (agRaw * ajio) / 100;
              const akValue = agRaw - ahValue - ajValue;
              const alPctVal = akValue > 2499 ? 0.18 : 0.05;
              const amValue = akValue * alPctVal;
              const anValue = akValue + amValue;

              // Col AA (27): COMPANY PROFIT MARGIN %
              const aaCell = addedRow.getCell(27);
              aaCell.value = {{
                formula: 'IF(Y' + R + '="","",ROUND((Y' + R + '*' + margin + ')/100,2))',
                result: profitNum
              }};
              aaCell.numFmt = '#,##0.00';

              // Col AD (30): SALE TP WITH TAX
              const adCell = addedRow.getCell(30);
              adCell.value = {{
                formula: 'IF(Y' + R + '="","",ROUND(Y' + R + '+AA' + R + ',2))',
                result: acNum
              }};
              adCell.numFmt = '#,##0.00';

              // Col AC (29): GST%
              const acCell = addedRow.getCell(29);
              acCell.value = {{
                formula: 'IF(AD' + R + '="","",ROUND((AD' + R + '*' + gst + ')/(100+' + gst + '),2))',
                result: abNum
              }};
              acCell.numFmt = '#,##0.00';

              // Col AB (28): SALE TP WITH OUT TAX
              const abCell = addedRow.getCell(28);
              abCell.value = {{
                formula: 'IF(AD' + R + '="","",ROUND(AD' + R + '-AC' + R + ',2))',
                result: aaNum
              }};
              abCell.numFmt = '#,##0.00';

              // Col AE (31): MRP (pure number from GoalSeek, no formula)
              const aeCell = addedRow.getCell(31);
              aeCell.value = aeRaw > 0 ? Math.round(aeRaw) : '';
              aeCell.numFmt = '#,##0';

              // Col AF (32): SALE DISCOUNT AMT
              const afCell = addedRow.getCell(32);
              afCell.value = {{
                formula: 'IF(AE' + R + '="","",ROUND((AE' + R + '*' + discount + ')/100,0))',
                result: Math.round(afRaw)
              }};
              afCell.numFmt = '#,##0';

              // Col AG (33): ASP (GROSS)
              const agCell = addedRow.getCell(33);
              agCell.value = {{
                formula: 'IF(AE' + R + '="","",AE' + R + '-AF' + R + ')',
                result: Math.round(agRaw)
              }};
              agCell.numFmt = '#,##0';

              // Col AH (34): GST%
              const ahCell = addedRow.getCell(34);
              ahCell.value = {{
                formula: 'IF(AG' + R + '="","",IF(AG' + R + '>2499,0.18,0.05))',
                result: ahPctVal
              }};
              ahCell.numFmt = '0%';

              // Col AI (35): GST amount
              const aiCell = addedRow.getCell(35);
              aiCell.value = {{
                formula: 'IF(AG' + R + '="","",ROUND(AG' + R + '-(AG' + R + '/(1+AH' + R + ')),2))',
                result: Math.round(ahValue * 100) / 100
              }};
              aiCell.numFmt = '#,##0.00';

              // Col AJ (36): NET SALE
              const ajCell = addedRow.getCell(36);
              ajCell.value = {{
                formula: 'IF(AG' + R + '="","",ROUND(AG' + R + '-AI' + R + ',2))',
                result: Math.round(aiValue * 100) / 100
              }};
              ajCell.numFmt = '#,##0.00';

              // Col AK (37): AJIO MARGIN 34%
              const akCell = addedRow.getCell(37);
              akCell.value = {{
                formula: 'IF(AG' + R + '="","",ROUND((AG' + R + '*' + ajio + ')/100,2))',
                result: Math.round(ajValue * 100) / 100
              }};
              akCell.numFmt = '#,##0.00';

              // Col AL (38): PURCHASE
              const alCell = addedRow.getCell(38);
              alCell.value = {{
                formula: 'IF(AG' + R + '="","",ROUND(AG' + R + '-AI' + R + '-AK' + R + ',2))',
                result: Math.round(akValue * 100) / 100
              }};
              alCell.numFmt = '#,##0.00';

              // Col AM (39): GST%2
              const amCell = addedRow.getCell(39);
              amCell.value = {{
                formula: 'IF(AL' + R + '="","",IF(AL' + R + '>2499,0.18,0.05))',
                result: alPctVal
              }};
              amCell.numFmt = '0%';

              // Col AN (40): GST2 amount
              const anCell = addedRow.getCell(40);
              anCell.value = {{
                formula: 'IF(AL' + R + '="","",ROUND(AL' + R + '*AM' + R + ',2))',
                result: Math.round(amValue * 100) / 100
              }};
              anCell.numFmt = '#,##0.00';

              // Col AO (41): BANK SETTLEMENT
              const aoCell = addedRow.getCell(41);
              aoCell.value = {{
                formula: 'IF(AL' + R + '="","",ROUND(AL' + R + '+AN' + R + ',2))',
                result: Math.round(anValue * 100) / 100
              }};
              aoCell.numFmt = '#,##0.00';
            }}
          }});

          // ── 4. Set column widths ──
          exportHeaders.forEach((h, i) => {{
            let max = String(h || '').length + 3;
            rows.forEach(r => {{
              const v = String(r[i] || '');
              const w = i === 2
                ? Math.min(Math.max(v.length + 3, 32), 80)
                : Math.min(Math.max(v.length + 3, 12), 28);
              if (w > max) max = w;
            }});
            ws.getColumn(i + 1).width = Math.max(max, 12);
          }});

          // ── 5. Style the header row (colors + borders + font + alignment) ──
          const headerRow = ws.getRow(1);
          headerRow.height = 20;

          const redFill = {{
            type: 'pattern',
            pattern: 'solid',
            fgColor: {{ argb: 'FFFDE2E2' }}
          }};
          const yellowFill = {{
            type: 'pattern',
            pattern: 'solid',
            fgColor: {{ argb: 'FFFFFF4B' }}
          }};
          const greenFill = {{
            type: 'pattern',
            pattern: 'solid',
            fgColor: {{ argb: 'FFDDF6E4' }}
          }};
          const thinBorder = {{
            top:    {{ style: 'thin', color: {{ argb: 'FF000000' }} }},
            bottom: {{ style: 'thin', color: {{ argb: 'FF000000' }} }},
            left:   {{ style: 'thin', color: {{ argb: 'FF000000' }} }},
            right:  {{ style: 'thin', color: {{ argb: 'FF000000' }} }}
          }};

          for (let C = 1; C <= exportHeaders.length; C++) {{
            const cell = headerRow.getCell(C);
            const fill = C <= 25 ? redFill : C === 26 ? yellowFill : greenFill;
            cell.fill = fill;
            cell.border = thinBorder;
            cell.font = {{
              name: 'Calibri',
              size: 11,
              bold: true,
              color: {{ argb: 'FF000000' }}
            }};
            cell.alignment = {{
              horizontal: 'center',
              vertical: 'middle',
              wrapText: false
            }};
          }}

          // ── 6. Apply % format to cell values that were percentages ──
          for (let R = 2; R <= rows.length + 1; R++) {{
            const dataRow = ws.getRow(R);
            for (let C = 1; C <= exportHeaders.length; C++) {{
              const cell = dataRow.getCell(C);
              if (C === 34 || C === 39) {{
                cell.numFmt = '0%';
                continue;
              }}
              const origVal = rows[R - 2][C - 1];
              const origStr = String(origVal || '').trim();
              if (origStr.endsWith('%')) {{
                cell.numFmt = '0%';
              }}
            }}
          }}

          // ── 7. Generate file & download ──
          const buffer = await workbook.xlsx.writeBuffer();
          const blob = new Blob([buffer], {{ type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' }});
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = safeName + '.xlsx';
          document.body.appendChild(a);
          a.click();
          document.body.removeChild(a);
          URL.revokeObjectURL(url);

          closeExportPopup();

        }} catch (err) {{
          console.error(err);
          showNotice('Export failed: ' + (err.message || 'Unknown error'), 'Export failed');
        }} finally {{
          confirmExportBtn.disabled = false;
          confirmExportBtn.textContent = 'Download File';
        }}
      }}

      window.exportToExcel = openExportPopup;
      window.confirmExport = exportExcel;
      window.cancelExport = closeExportPopup;
      
      window.openClearPopup = () => clearPopup.classList.add('open');
      window.closeClearPopup = () => clearPopup.classList.remove('open');

      if (prevBtn && nextBtn) {{
        prevBtn.addEventListener('click', () => {{
          if (currentPage > 1) {{
            currentPage--;
            renderPage();
          }}
        }});
        nextBtn.addEventListener('click', () => {{
          const totalPages = Math.ceil(tableData.length / pageSize);
          if (currentPage < totalPages) {{
            currentPage++;
            renderPage();
          }}
        }});
      }}


      if (marginInput) {{
        marginInput.addEventListener('input', updateCompanyProfit);
        marginInput.addEventListener('input', updateSingleMode);
      }}
      if (gstInput) {{
        gstInput.addEventListener('input', updateCompanyProfit);
        gstInput.addEventListener('input', updateSingleMode);
      }}
      if (discountInput) {{
        discountInput.addEventListener('input', updateCompanyProfit);
        discountInput.addEventListener('input', updateSingleMode);
      }}
      if (ajioInput) {{
        ajioInput.addEventListener('input', updateCompanyProfit);
        ajioInput.addEventListener('input', updateSingleMode);
      }}

      if (openExportPopupBtn) {{
        openExportPopupBtn.addEventListener('click', openExportPopup);
      }}
      if (clearDataBtn) {{
        clearDataBtn.addEventListener('click', () => {{
          if (clearPopup) clearPopup.classList.add('open');
        }});
      }}
      if (cancelClearBtn) {{
        cancelClearBtn.addEventListener('click', () => {{
          if (clearPopup) clearPopup.classList.remove('open');
        }});
      }}
      if (confirmClearBtn) {{
        confirmClearBtn.addEventListener('click', () => {{
          window.location.href = '/';
        }});
      }}
      if (clearPopup) {{
        clearPopup.addEventListener('click', (event) => {{
          if (event.target === clearPopup) {{
            clearPopup.classList.remove('open');
          }}
        }});
      }}
      if (cancelExportBtn) {{
        cancelExportBtn.addEventListener('click', closeExportPopup);
      }}
      if (confirmExportBtn) {{
        confirmExportBtn.addEventListener('click', exportExcel);
      }}
      if (exportPopup) {{
        exportPopup.addEventListener('click', (event) => {{
          if (event.target === exportPopup) {{
            closeExportPopup();
          }}
        }});
      }}
      if (exportFileNameInput) {{
        exportFileNameInput.addEventListener('keydown', (event) => {{
          if (event.key === 'Enter') {{
            event.preventDefault();
            exportExcel();
          }}
        }});
      }}

      // Chatbot Logic
      const botBubble = document.getElementById('bot-bubble');
      const botWindow = document.getElementById('bot-window');
      const botMessages = document.getElementById('bot-messages');
      const botInput = document.getElementById('bot-input');
      
      let chatState = {{
        userName: null,
        step: 0,
        data: {{}}
      }};

      const steps = [
        {{ key: 'y', label: 'PURCHASE TP (Y)', prompt: (name) => `Alright ${{name}}! First, please tell me your PURCHASE TP (Y): 💰` }},
        {{ key: 'pgst', label: 'PURCHASE GST %', prompt: (name) => `Thank you ${{name}}! Now, what is the PURCHASE GST %? 💸` }},
        {{ key: 'margin', label: 'COMPANY PROFIT MARGIN %', prompt: (name) => `Got it ${{name}}! What is the COMPANY PROFIT MARGIN %? 📈` }},
        {{ key: 'sgst', label: 'SALE GST %', prompt: (name) => `Great ${{name}}! Now tell me the SALE GST %: 🏛️` }},
        {{ key: 'discount', label: 'SALE DISCOUNT AMT %', prompt: (name) => `Almost there ${{name}}! What is the SALE DISCOUNT AMT %? 🏷️` }},
        {{ key: 'ajio', label: 'AJIO MARGIN %', prompt: (name) => `Last step ${{name}}! What is the AJIO MARGIN %? 🤝` }}
      ];

      const randomGreetings = [
        "Hii I'am AJ your friend how can i help you",
        "Need help with calculations? ⚡",
        "Type 'calculate' to start a step-by-step guide! 📖",
        "I can help you find the Bank Settlement! 💰",
        "Don't forget to export your results! ✅"
      ];

      function typeWriter(text, i, cb) {{
        if (i < text.length) {{
          botBubble.textContent += text.charAt(i);
          setTimeout(() => typeWriter(text, i + 1, cb), 50);
        }} else if (cb) {{
          cb();
        }}
      }}

      function rotateGreetings() {{
        if (botWindow.classList.contains('open')) {{
            botBubble.classList.remove('show');
            return;
        }}
        const randomMsg = randomGreetings[Math.floor(Math.random() * randomGreetings.length)];
        
        botBubble.classList.remove('show');
        setTimeout(() => {{
            botBubble.textContent = '';
            botBubble.classList.add('show');
            typeWriter(randomMsg, 0, () => {{
                setTimeout(() => botBubble.classList.remove('show'), 6000);
            }});
        }}, 500);
      }}

      setInterval(rotateGreetings, 12000);
      setTimeout(rotateGreetings, 2000);

      window.toggleChat = () => {{
        botWindow.classList.toggle('open');
        if (botWindow.classList.contains('open')) botBubble.classList.remove('show');
      }};

      window.sendMessage = async () => {{
        const text = botInput.value.trim();
        if (!text) return;

        appendMessage('user', text);
        botInput.value = '';

        // Handle User Name first
        if (!chatState.userName) {{
            chatState.userName = text;
            appendMessage('bot', `Nice to meet you, ${{chatState.userName}}! 👋 How can I help you today? You can type "calculate" to start a guided cost check! ✨`);
            return;
        }}

        // Check if user wants to start or continue calculation
        const numVal = parseFloat(text);
        
        if (text.toLowerCase().includes('calculate') || text.toLowerCase().includes('tp')) {{
            chatState.step = 0;
            chatState.data = {{}};
            appendMessage('bot', steps[0].prompt(chatState.userName));
            chatState.step = 1;
            return;
        }}

        if (chatState.step > 0 && chatState.step <= steps.length) {{
            if (isNaN(numVal)) {{
                appendMessage('bot', `Kripya ek sahi number batayein ${{steps[chatState.step-1].label}} ke liye.`);
                return;
            }}
            
            chatState.data[steps[chatState.step-1].key] = numVal;
            
            if (chatState.step < steps.length) {{
                appendMessage('bot', steps[chatState.step].prompt(chatState.userName));
                chatState.step++;
            }} else {{
                // Final Calculation
                const result = calculateFinal(chatState.data);
                appendMessage('bot', `Alright ${{chatState.userName}}, your calculation is ready! ✨\n\n` + 
                    `🛒 SALE TP WITH TAX (AD): ₹${{result.ad}}\n` +
                    `🏷️ MRP (AE): ₹${{result.ae}}\n` +
                    `📦 ASP (GROSS) (AG): ₹${{result.ag}}\n` +
                    `✅ BANK SETTLEMENT (AO): ₹${{result.ao}}\n\n` +
                    `Is there anything else I can help you with?`);
                chatState.step = 0;
            }}
            return;
        }}

        // Normal AI Chat
        try {{
          const response = await fetch('/api/chat', {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify({{
              message: text,
              userName: chatState.userName
            }})
          }});

          if (!response.ok) throw new Error('API Error');
          const data = await response.json();
          const botText = data.reply;
          appendMessage('bot', botText);
        }} catch (e) {{
          appendMessage('bot', "I'am sorry, I'am having trouble connecting right now. 😔");
        }}
      }};

      function calculateFinal(d) {{
        const valAA = Math.round(((d.y * d.margin) / 100) * 100) / 100;
        const valAD = Math.round((d.y + valAA) * 100) / 100;
        
        const aeRaw = (valAD <= 1606) ? (valAD / 0.22505) : 
                      (valAD <= 2620) ? (valAD / 0.186490678) : 
                      (valAD / 0.209579661);

        const afRaw = (aeRaw * d.discount) / 100;
        const agRaw = aeRaw - afRaw;
        const ahPctVal = agRaw > 2499 ? 0.18 : 0.05;
        const ahValue = agRaw - (agRaw / (1 + ahPctVal));
        const aiValue = agRaw - ahValue;
        const ajValue = (agRaw * d.ajio) / 100;
        const akValue = agRaw - ahValue - ajValue;
        const alPctVal = akValue > 2499 ? 0.18 : 0.05;
        const amValue = akValue * alPctVal;
        const anValue = akValue + amValue;
        
        return {{ 
          ad: valAD.toFixed(2),
          ae: Math.round(aeRaw).toFixed(0),
          ag: Math.round(agRaw).toFixed(0),
          ao: anValue.toFixed(2) 
        }};
      }}

      function appendMessage(type, text) {{
        const div = document.createElement('div');
        div.className = `msg ${{type}}`;
        div.style.whiteSpace = 'pre-wrap';
        div.textContent = text;
        botMessages.appendChild(div);
        botMessages.scrollTop = botMessages.scrollHeight;
      }}

      renderPage();
    }});
  </script>
</body>
</html>
"""
    return page.encode("utf-8")


@app.get("/sample-excel")
def download_sample_excel() -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Template"
    
    # A1: Item Details
    ws["A1"] = "Item Details"
    ws["A1"].font = Font(size=14)
    
    # Row 2: Headers
    headers = [
        "No.", "Seller Name", "Item Name", "Category", "Brand", "Type", 
        "ASIN Number", "SKU Number", "HSN Number", "MRP Price", "Item Color", 
        "Weight", "Weight Unit", "Length", "Length Unit", "Width", 
        "Width Unit", "Height", "Height Unit", "Channel Price", 
        "Purchase Margin(%)", "Tax", "Purchase Cost", "Purchase Tax", 
        "Final Purchase Cost"
    ]
    ws.append(headers)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = "attachment; filename=ajio_sample_template.xlsx"
    return response


@app.get("/")
def index() -> Response:
    return Response(render_page(), mimetype="text/html")


@app.get("/logo.png")
def serve_logo() -> Response:
    return send_from_directory(BASE_DIR, "logo.png")


@app.get("/chatbot.png")
def serve_chatbot() -> Response:
    return send_from_directory(BASE_DIR, "chatbot.png")


@app.post("/api/chat")
def chat_route() -> Response:
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        return jsonify({"error": "AI chat is not configured."}), 503

    payload = request.get_json(force=True, silent=True) or {}
    message = str(payload.get("message", "")).strip()
    user_name = str(payload.get("userName", "friend")).strip() or "friend"

    if not message:
        return jsonify({"error": "Message is required."}), 400

    gemini_payload = {
        "system_instruction": {
            "parts": [
                {
                    "text": f"""You are AJ, the user's "Dashboard Friend" and business assistant for this AJIO Cost Sheet.
Your tone is extremely friendly, professional, and helpful. Always address the user as "{user_name}".

CORE CAPABILITIES:
1. Simple Math: Perform basic arithmetic (+, -, *, /).
2. Percentage Logic: Handle queries like "X% of Y", "percentage of X", etc.
3. Advanced Business Math:
   - Reverse GST: Extract base price from a tax-inclusive price (Formula: Price / (1 + TaxRate)).
   - Profit Margin vs Markup: Calculate required selling price for a target margin.
   - Discount Stacking: Calculate final price after sequential discounts (e.g., 20% + 10%).
   - Break-even: Calculate minimum price to cover costs.
4. General Support: Answer questions about the AJIO dashboard or cost calculations.

GUIDELINES:
- Respond in "Hinglish" (a natural mix of Hindi and English).
- For math queries, clearly state the result clearly and briefly show the formula or steps used.
- If the user seems confused, offer to start the guided calculation by telling them to type "calculate".
- Keep responses concise but warm."""
                }
            ]
        },
        "contents": [{"parts": [{"text": message}]}],
    }

    try:
        gemini_response = requests.post(
            GEMINI_API_URL,
            params={"key": api_key},
            json=gemini_payload,
            timeout=30,
        )
        gemini_response.raise_for_status()
        data = gemini_response.json()
        reply = data["candidates"][0]["content"]["parts"][0]["text"]
        return jsonify({"reply": reply})
    except (KeyError, IndexError, requests.RequestException) as exc:
        app.logger.exception("Gemini chat request failed: %s", exc)
        return jsonify({"error": "AI chat request failed."}), 502


@app.post("/export")
def export_excel_route() -> Response:
    try:
        payload = request.get_json(force=True, silent=True)
        if not payload or not isinstance(payload, dict):
            app.logger.error("Export called with invalid or empty JSON payload")
            return jsonify({"error": "Invalid request payload."}), 400
        filename = sanitize_filename(str(payload.get("filename", "ajio_export")))
        headers = [ILLEGAL_CHARACTERS_RE.sub("", str(cell)) for cell in payload.get("headers", [])]
        rows = payload.get("rows", []) or []
        if not headers:
            return jsonify({"error": "No headers provided."}), 400
        workbook_bytes = build_export_workbook(headers, rows)
    except Exception as exc:  # noqa: BLE001
        app.logger.exception("Excel export failed: %s", exc)
        return jsonify({"error": f"Excel export failed: {exc}"}), 500
    response = Response(
        workbook_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


@app.post("/")
def upload_excel_route() -> Response:
    try:
        upload = request.files.get("file")
        if upload is None or not upload.filename:
            raise ValueError("Koi file receive nahi hui.")

        filename = upload.filename
        file_bytes = upload.read()
        extension = os.path.splitext(filename)[1].lower()

        try:
            if extension == ".xls":
                headers, rows = parse_xls(file_bytes)
            else:
                headers, rows = parse_excel(file_bytes)
        except Exception:
            headers, rows = parse_with_calamine(file_bytes)

        # Apply Purchase GST logic if toggle is on
        if request.form.get("purchase_gst_toggle") == "on":
            for row in rows:
                if len(row) > 24: # Column Y is index 24
                    try:
                        # Ensure Y is a valid number
                        y_str = str(row[24]).replace(",", "").strip()
                        if not y_str: continue
                        val_y = float(y_str)
                        
                        val_v = 5
                        val_x = round(val_y * 5 / 105, 2)
                        val_w = round(val_y - val_x, 2)
                        
                        row[21] = str(val_v)  # V
                        row[23] = f"{val_x:.2f}" # X
                        row[22] = f"{val_w:.2f}" # W
                    except (ValueError, TypeError, IndexError):
                        continue

        page = render_page(render_table(headers, rows), f"{filename} successfully loaded.", default_tab="bulk")
        return Response(page, mimetype="text/html")
    except Exception as exc:  # noqa: BLE001
        error_page = render_page(message=f"File process nahi hui: {exc}")
        return Response(error_page, mimetype="text/html", status=400)


def run() -> None:
    port = int(os.environ.get("PORT", "8000"))
    app.run(host="0.0.0.0", port=port, debug=False)


if __name__ == "__main__":
    run()
